import xlsxwriter
from datetime import datetime
from flask import session
from models import cvValues, cvTable
from dbsetup import db
import openpyxl
from openpyxl.chart import LineChart,Reference

# wb = openpyxl.Workbook()
# sheet = wb.active
# from app import getDBElementWithId

def getDBElementWithId(table_name, id):
    # with app.app_context():
    if id:
        output_element = db.session.query(table_name).filter_by(id=id).first()
        return output_element
    else:
        return None
    


def createcvOpening_gas(itemCase_list,fluid_types,items,header_details):
    print(f'INSIDE GASOPENING')

    data = [10,20,30,40,50,60,70,80,90,100]
    workbook = xlsxwriter.Workbook(f'E:\Sizing_Reports\cvsizingcalculation.xlsx')   
    print(f'itemCasessssssssss',itemCase_list)
    count = 1;f_cnt = 0
    print(f'FLUID {fluid_types}')
    for itemCase in itemCase_list:
            
            cv_graph_values = []
            open_cv = [10,20,30,40,50,60,70,80,90,100]
            item = items[f_cnt]
            fluid_type = fluid_types[f_cnt]
            worksheet = workbook.add_worksheet()



            worksheet.set_paper(9)
            worksheet.fit_to_pages(2, 1)
            worksheet.print_area('A2:M72')
            worksheet.set_margins(0.36, 0.2, 0.4, 0.01)
            # worksheet.setRowHeight(2, 50)
            # worksheet.set_row(9, 10)
            worksheet.set_header(margin=0.2)
            worksheet.set_footer(margin=0.28)


            bold = workbook.add_format(
                {'bold': True, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1, 'font': 'Arial', 'font_size': 9, 'border': 7, 'bg_color': '#EFF5F5'})
            bold1 = workbook.add_format(
                {'bold': True, 'bottom': 1, 'top': 1, 'left': 1, 'border': 7, 'bg_color': '#EFF5F5'})
            boldc = workbook.add_format(
                {'align': 'center', 'bold': True, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1,'font': 'Arial', 'font_size': 9, 'border': 7,
                'bg_color': '#EFF5F5'})
            br = workbook.add_format({'bottom': 1, 'top': 1, 'right': 1, 'border': 7})
            h1 = workbook.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'font': 'Arial', 'bottom': 1, 'top': 1, 'right': 1, 'left': 1})
            h2 = workbook.add_format({'bold': 0, 'font_size': 10, 'font': 'Arial'})
            h3 = workbook.add_format({'bold': 0, 'font_size': 9, 'font': 'Arial'})
            f1 = workbook.add_format({'bold': 0, 'font_size': 9, 'font': 'Arial', 'align': 'right'})

            cell_format1 = workbook.add_format({'bottom': 1, 'top': 1, 'font': 'Arial', 'font_size': 9, 'border': 7, 'right': 0, 'left': 0})
            cell_formatL = workbook.add_format({'bottom': 1, 'top': 1, 'font': 'Arial', 'font_size': 9,  'right': 1,'border': 7, 'left': 0 })
            cell_format = workbook.add_format(
                {'bold': 0, 'font_color': 'black', 'font': 'Arial', 'font_size': 9, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1,
                'border': 7,'align': 'left'})

            cell_format2 = workbook.add_format(
                {'bold': 0, 'align': 'center', 'font_color': 'black', 'font': 'Arial', 'font_size': 9, 'bottom': 1, 'top': 1,
                'right': 1, 'left': 1, 'border': 7})

            cell_format3 = workbook.add_format(
            {'bold': True, 'align': 'center', 'font_color': 'black', 'font': 'Arial', 'font_size': 12, 'bottom': 1, 'top': 1,
                'right': 1, 'left': 1, 'border': 7})
            h3 = workbook.add_format({'bold': 0, 'font_size': 9, 'font': 'Arial'})

            column_width = [{'name': 'A1:A73', 'size': 5}, {'name': 'B1:B73', 'size': 16.57}, {'name': 'C1:C73', 'size': 11},
                            {'name': 'D1:D73', 'size': 12},
                            {'name': 'E1:E73', 'size': 7.5}, {'name': 'F1:F73', 'size': 7}, {'name': 'G1:G73', 'size': 5},
                            {'name': 'H1:H73', 'size': 9.5},
                            {'name': 'I1:I73', 'size': 13}, {'name': 'J1:J73', 'size': 7.5}, {'name': 'K1:K73', 'size': 7},
                            {'name': 'L1:L73', 'size': 6.5}, {'name': 'M1:M73', 'size': 8}]

            for i in column_width:
                worksheet.set_column(i['name'], i['size'])

            # worksheet.set_column('A1:A73', 5)
            # worksheet.set_column('B1:B73', 24)
            # worksheet.set_column('C1:C73', 10)
            # worksheet.set_column('D1:D73', 12)
            # worksheet.set_column('E1:E73', 7)
            # worksheet.set_column('F1:F73', 5)
            # worksheet.set_column('G1:G73', 12)
            # worksheet.set_column('H1:H73', 12)
            # worksheet.set_column('I1:I73', 6)
            # worksheet.set_column('J1:J73', 6)
            # worksheet.set_column('K1:K73', 6)
            # worksheet.set_column('L1:L73', 6)

            # worksheet.insert_image('A4', 'logo.png', {'x_scale': 0.2, 'y_scale': 0.2})
            worksheet.merge_range('A3:B7', "", cell_format)
            worksheet.insert_image('A4', 'newlogo.png',{'x_scale': 0.18, 'y_scale': 0.18})
            worksheet.write('A1', '', )
            worksheet.merge_range('A2:M2', 'Control Valve Sizing Calculation Sheet', cell_format3)
            worksheet.merge_range('A8:B8', "Application", cell_format)
            worksheet.write('A9:B9', 'Fluid State / Name', cell_format)
            worksheet.write('A10', 'AA', boldc)
            worksheet.write('A11', '1', cell_format2)
            worksheet.write('A12', '2', cell_format2)
            worksheet.write('A13', '3', cell_format2)
            worksheet.write('A14', '4', cell_format2)
            worksheet.write('A15', '5', cell_format2)
            worksheet.write('A16', '6', cell_format2)
            worksheet.write('A17', '7', cell_format2)
            worksheet.write('A18', '8', cell_format2)
            worksheet.write('A19', '9', cell_format2)
            worksheet.write('A20', '10', cell_format2)
            worksheet.write('A21', '11', cell_format2)
            worksheet.write('A22', '12', cell_format2)
            worksheet.write('A23', '13', cell_format2)
            worksheet.write('A24', '14', cell_format2)
            worksheet.write('A25', '15', cell_format2)
            worksheet.write('A26', '16', cell_format2)
            worksheet.write('A27', '17', cell_format2)
            worksheet.write('A28', '18', cell_format2)
            worksheet.write('A29', '19', cell_format2)
            worksheet.write('A30', '20', cell_format2)
            worksheet.write('A31', '21', cell_format2)
            worksheet.write('A32', '22', cell_format2)
            worksheet.write('A33', '23', cell_format2)
            worksheet.write('A34', '24', cell_format2)
            worksheet.write('A35', '25', cell_format2)
            worksheet.write('A36', '26', cell_format2)
            worksheet.write('A37', '27', cell_format2)
            if fluid_type == 'Gas':
                worksheet.write('A38', '28', cell_format2)
                worksheet.write('A39', '29', cell_format2)
                worksheet.write('A40', '30', cell_format2)
                worksheet.write('A41', '31', cell_format2)
                worksheet.write('A42', '32', cell_format2)
                # worksheet.write('A43', '33', cell_format2)
                # worksheet.write('A44', '34', cell_format2)
                # worksheet.write('A45', '35', cell_format2)
                # worksheet.write('A46', '36', cell_format2)
                # worksheet.write('A47', '37', cell_format2)
                # worksheet.write('A48', '38', cell_format2)



            worksheet.write('B2', '')
            worksheet.write('B3', '')
            if fluid_type == 'Gas':
                worksheet.merge_range("B10:C10", 'Valve Sizing Calculation', bold)
                worksheet.merge_range("B11:C11", 'Flow Rate', cell_format)
                worksheet.merge_range("B12:C12", 'Inlet Pressure', cell_format)
                worksheet.merge_range("B13:C13", 'Outlet Pressure', cell_format)
                worksheet.merge_range("B14:C14", 'Inlet Temperature', cell_format)
                worksheet.merge_range("B15:C15", 'Molecular Weight', cell_format)
                worksheet.merge_range("B16:C16", 'Spec. Heat Ratio', cell_format)
                worksheet.merge_range("B17:C17", 'Inlet compressibility factor', cell_format)
                worksheet.merge_range("B18:C18", 'Pressure drop ratio factor', cell_format)
                worksheet.merge_range("B19:C19", 'Calculated Cv', cell_format)
                worksheet.merge_range("B20:C20", 'Percentage opening / Degree', cell_format)
                worksheet.merge_range("B21:C21", 'Noise level at 1m distance', cell_format)
                worksheet.merge_range("B22:C22", 'Ratio of Pressure drop (dp/P1)', cell_format)
                worksheet.merge_range("B23:C23", 'Choked pressure drop', cell_format)
                worksheet.merge_range("B24:C24", 'Expansion Ratio ', cell_format)
                worksheet.merge_range("B25:C25", 'Pressure drop ratio factor', cell_format)
                worksheet.merge_range("B26:C26", 'Ratio of specific heat factor', cell_format)
                worksheet.merge_range("B27:C27", 'Piping Geometry factor', cell_format1)
                worksheet.merge_range("B28:C28", 'Application Ratio', cell_format1)
                worksheet.merge_range("B29:C29", "Reynold's Number", cell_format1)
                # worksheet.merge_range("B30:C30", 'Kc', cell_format1)
                worksheet.merge_range("B30:C30", 'Valve Diameter', cell_format1)
                worksheet.merge_range("B31:C31", 'Inlet Pipe Velocity', cell_format1)
                # worksheet.merge_range("B33:C33", 'Noise level at 1m distance', cell_format1)
                worksheet.merge_range("B32:C32", 'Mach No. Upstream', cell_format1)
                worksheet.merge_range("B33:C33", 'Mach No. Downstream', cell_format1)
                worksheet.merge_range("B34:C34", 'Mach No. Valve Outlet', cell_format1)
                worksheet.merge_range("B35:C35", 'Sonic Velocity Upstream', cell_format1)
                worksheet.merge_range("B36:C36", 'Sonic Velocity Downstream', cell_format1)
                worksheet.merge_range("B37:C37", 'Sonic velocity Valve Outlet', cell_format1)
                worksheet.merge_range("B38:C38", 'Outlet Density', cell_format1)
                worksheet.merge_range("B39:C39", 'Trim Exit Velocity', cell_format1)
                worksheet.merge_range("B40:C40", 'Power Lvl(kw)', cell_format1) 
                worksheet.merge_range("B41:C41", 'Required Stages', cell_format1)
                worksheet.merge_range("B42:C42", 'Warnings!!', cell_format1)
           
            elif fluid_type == 'Liquid':
                worksheet.merge_range("B10:C10", 'Valve Sizing Calculation', bold)
                worksheet.merge_range("B11:C11", 'Flow Rate', cell_format)
                worksheet.merge_range("B12:C12", 'Inlet Pressure', cell_format)
                worksheet.merge_range("B13:C13", 'Outlet Pressure', cell_format)
                worksheet.merge_range("B14:C14", 'Inlet Temperature', cell_format)
                worksheet.merge_range("B15:C15", 'Specific Gravity', cell_format)
                worksheet.merge_range("B16:C16", 'Kinematic Viscosity', cell_format)
                worksheet.merge_range("B17:C17", 'Vapor Pressure', cell_format)
                worksheet.merge_range("B18:C18", 'Liquid Pr. Recovery factor', cell_format)
                worksheet.merge_range("B19:C19", 'Calculated Cv', cell_format)
                worksheet.merge_range("B20:C20", 'Percentage opening / Degree', cell_format)
                worksheet.merge_range("B21:C21", 'Noise level at 1m distance', cell_format)
                worksheet.merge_range("B22:C22", 'Differential Pressure Drop', cell_format)
                worksheet.merge_range("B23:C23", 'Choked pressure drop', cell_format)
                worksheet.merge_range("B24:C24", 'Ff', cell_format1)
                worksheet.merge_range("B25:C25", 'Fp', cell_format1)
                worksheet.merge_range("B26:C26", 'FLp', cell_format1)
                worksheet.merge_range("B27:C27", 'Kc', cell_format1)
                worksheet.merge_range("B28:C28", 'Ar', cell_format1)
                worksheet.merge_range("B29:C29", 'Re', cell_format1)
                worksheet.merge_range("B30:C30", 'Valve Diameter', cell_format1)        
                worksheet.merge_range("B31:C31", 'Inlet pipe velocity', cell_format1)
                worksheet.merge_range("B32:C32", 'Outlet pipe velocity', cell_format1)
                worksheet.merge_range("B33:C33", 'Valve Velocity', cell_format1)
                worksheet.merge_range("B34:C34", 'Trim Exit velocity', cell_format1)
                worksheet.merge_range("B35:C35", 'Power Lvl', cell_format1)
                worksheet.merge_range("B36:C36", 'Required Stages', cell_format1)
                worksheet.merge_range("B37:C37", 'Warnings!!', cell_format1)

 


            worksheet.write('C1', '')
            # worksheet.write('C2', '')
            worksheet.write('C3', 'Customer', cell_format)
            worksheet.write('C4', 'Project', cell_format)
            worksheet.write('C5', 'End User', cell_format)
            worksheet.write('C6', 'RFQ No', cell_format)
            worksheet.write('C7', 'PO No', cell_format)
            worksheet.merge_range('C8:M8', '', cell_format)
            worksheet.merge_range('C9:F9', '', cell_format)

            # worksheet.write('D2', '')
            worksheet.merge_range('D3:H3', '', cell_format)
            worksheet.merge_range('D4:H4', '', cell_format)
            worksheet.merge_range('D5:H5', '', cell_format)
            worksheet.merge_range('D6:H6', '', cell_format)
            worksheet.merge_range('D7:H7', '', cell_format)

            worksheet.write('D10', 'Units', bold)
            worksheet.write('D11', '', cell_format)
            worksheet.write('D12', '', cell_format)
            worksheet.write('D13', '', cell_format)
            worksheet.write('D14', '', cell_format)
            worksheet.write('D15', '', cell_format)
            worksheet.write('D16', '', cell_format)
            worksheet.write('D17', '', cell_format)
            worksheet.write('D18', '', cell_format)
            worksheet.write('D19', '', cell_format)
            worksheet.write('D20', '', cell_format)
            worksheet.write('D21', '', cell_format)
            worksheet.write('D22', '', cell_format)
            worksheet.write('D23', '', cell_format)
            worksheet.write('D24', '', cell_format)
            worksheet.write('D25', '', cell_format)
            worksheet.write('D26', '', cell_format)
            worksheet.write('D27', '', cell_format)
            worksheet.write('D28', '', cell_format)
            worksheet.write('D29', '', cell_format)
            worksheet.write('D30', '', cell_format)
            worksheet.write('D31', '', cell_format)
            worksheet.write('D32', '', cell_format)
            worksheet.write('D33', '', cell_format)
            worksheet.write('D34', '', cell_format)
            worksheet.write('D35', '', cell_format)
            worksheet.write('D36', '', cell_format)
            worksheet.write('D37', '', cell_format)
            worksheet.write('D38', '', cell_format)
            
            worksheet.write('D40', '', cell_format)
            worksheet.write('D41', '', cell_format)
            worksheet.write('D42', '', cell_format)
            # worksheet.write('D43', '', cell_format)
            # worksheet.write('D44', '', cell_format)
            # worksheet.write('D45', '', cell_format)
            # worksheet.write('D46', '', cell_format)
            # worksheet.write('D47', '', cell_format)


            if fluid_type == 'Liquid':
                
                worksheet.write('D11', item.flowrate_unit, cell_format)
                worksheet.write('D12', item.inpres_unit, cell_format)
                worksheet.write('D13', item.outpres_unit, cell_format)
                worksheet.write('D14', f'°{item.intemp_unit}', cell_format)
                worksheet.write('D17', item.vaporpres_unit, cell_format)
                worksheet.write('D30', item.valvesize_unit, cell_format)
            elif fluid_type == 'Gas':
                worksheet.write('D39', 'kW',cell_format )
                worksheet.write('D11', item.flowrate_unit, cell_format)
                worksheet.write('D12', item.inpres_unit, cell_format)
                worksheet.write('D13', item.outpres_unit, cell_format)
                worksheet.write('D14', f'°{item.intemp_unit}', cell_format)
                worksheet.write('D17', 'Z', cell_format)
                worksheet.write('D18', 'XT', cell_format)
                worksheet.write('D19', 'Cv', cell_format)
                worksheet.write('D20', '%', cell_format)
                worksheet.write('D21', 'dBA', cell_format)
                worksheet.write('D22', 'x', cell_format)
                worksheet.write('D24', 'Y', cell_format)
                worksheet.write('D25', 'Xtp', cell_format)
                worksheet.write('D26', 'Fk', cell_format)
                worksheet.write('D27', 'Fp', cell_format)
                worksheet.write('D28', 'Ar', cell_format)
                worksheet.write('D29', 'Re', cell_format)
                worksheet.write('D30', item.valvesize_unit, cell_format)
                worksheet.write('D31', 'm/s', cell_format)
                worksheet.write('D32', 'Mach', cell_format)
                worksheet.write('D33', 'Mach', cell_format)
                worksheet.write('D34', 'Mach', cell_format)
                worksheet.write('D35', 'm/s', cell_format)
                worksheet.write('D36', 'm/s', cell_format)
                worksheet.write('D37', 'm/s', cell_format)
                worksheet.write('D38', 'kg/m3', cell_format)
                worksheet.write('D39', 'kW', cell_format)   




            print(f'HEADER DETAILS, {header_details}')
            worksheet.write('D3', header_details[f_cnt][0], cell_format1)
            worksheet.write('D4', header_details[f_cnt][1], cell_format1)
            worksheet.write('D5', header_details[f_cnt][2], cell_format1)
            worksheet.write('D6', header_details[f_cnt][3], cell_format1)
            worksheet.write('D7', header_details[f_cnt][4], cell_format1)

            



           
            # worksheet.write('J6', str(case_data[i][0][14]), cell_format1) 
            # worksheet.write('J7', str(case_data[i][0][15]), cell_format1)
            # worksheet.write('D27', f"{str(case_data[i][0][18])} {units[i][-1]}", cell_format1)
            # worksheet.write('D28', f"{str(case_data[i][0][19])} {units[i][-2]}", cell_format1)


            worksheet.merge_range("E10:F10", 'Case 1', bold)
            case_cell = ["E:F","G:H","I:I","J:K","L:M"]
            print(f'HHHHHHHHHSGGGGGGGGGGGG ')
            if fluid_type == 'Gas' :
                cv_graph = 0; cv_graph_values = [];case_cv_final =[];case_cv_finalPercent = []
                for i in range(5):
                    st_cell,end_cell = case_cell[i].split(":") 
                    print(f'ITEMCASESS {itemCase}')
                    if i < len(itemCase): 

                        item_case = itemCase[i]   
                        case_cv_final.append(item_case.calculatedCv)
                        case_cv_finalPercent.append(item_case.openingPercentage)  
                        if item_case.valveDiaId:
                            # case_cv_values
                            cv_graph = 1

                            cv_element = getDBElementWithId(cvTable, item_case.valveDiaId)
                            # db.session.query(cvValues).filter_by(cv=cv_element, coeff='Cv').first()
                            item_case_cvs = db.session.query(cvValues).filter_by(cv=cv_element, coeff='Cv').first()
                            print(f'itemcasecvssssddvv {item_case_cvs}')
                            cv_graph_values = [item_case_cvs.one, item_case_cvs.two, item_case_cvs.three,item_case_cvs.four,item_case_cvs.five,item_case_cvs.six,item_case_cvs.seven,item_case_cvs.eight,item_case_cvs.nine,item_case_cvs.ten]
                        else:
                            print(f'Select Valve to view graph')   
                        if i == 2:
                            worksheet.write(f"{st_cell}{11}", item_case.flowrate, cell_format)
                            worksheet.write(f"{st_cell}{12}", item_case.inletPressure, cell_format)
                            worksheet.write(f"{st_cell}{13}:", item_case.outletPressure, cell_format)
                            worksheet.write(f"{st_cell}{14}:", item_case.inletTemp, cell_format)
                            worksheet.write(f"{st_cell}{15}:", item_case.molecularWeight, cell_format)
                            worksheet.write(f"{st_cell}{16}:", item_case.specificHeatRatio, cell_format)
                            worksheet.write(f"{st_cell}{17}:", item_case.compressibility, cell_format)
                            worksheet.write(f"{st_cell}{18}:", item_case.xt, cell_format)
                            worksheet.write(f"{st_cell}{19}:", item_case.calculatedCv, cell_format)
                            worksheet.write(f"{st_cell}{20}:", item_case.openingPercentage, cell_format)
                            worksheet.write(f"{st_cell}{21}:", item_case.spl, cell_format)
                            worksheet.write(f"{st_cell}{22}:", item_case.x_delp, cell_format)
                            worksheet.write(f"{st_cell}{23}:", item_case.chokedDrop, cell_format)
                            worksheet.write(f"{st_cell}{24}:", item_case.y_expansion, cell_format)
                            worksheet.write(f"{st_cell}{25}:", item_case.xtp, cell_format)
                            worksheet.write(f"{st_cell}{26}:", item_case.fk, cell_format)
                            worksheet.write(f"{st_cell}{27}:", item_case.Fp, cell_format)
                            worksheet.write(f"{st_cell}{28}:", item_case.ar, cell_format)
                            worksheet.write(f"{st_cell}{29}:", item_case.reNumber, cell_format)
                            worksheet.write(f"{st_cell}{30}:", item_case.valveSize, cell_format)
                            worksheet.write(f"{st_cell}{31}:", item_case.pipeInVel, cell_format)
                            worksheet.write(f"{st_cell}{32}:", item_case.machNoUp, cell_format)
                            worksheet.write(f"{st_cell}{33}:", item_case.machNoDown, cell_format)
                            worksheet.write(f"{st_cell}{34}:", item_case.machNoValve, cell_format)
                            worksheet.write(f"{st_cell}{35}:", item_case.sonicVelUp, cell_format)
                            worksheet.write(f"{st_cell}{36}:", item_case.sonicVelDown, cell_format)
                            worksheet.write(f"{st_cell}{37}:", item_case.sonicVelValve, cell_format)
                            worksheet.write(f"{st_cell}{38}:", item_case.outletDensity, cell_format)
                            worksheet.write(f"{st_cell}{39}:", item_case.tex, cell_format)
                            worksheet.write(f"{st_cell}{40}:", item_case.powerLevel, cell_format)
                            worksheet.write(f"{st_cell}{41}:", item_case.requiredStages, cell_format)
                            worksheet.write(f"{st_cell}{42}:", '', cell_format)
                           
                                 

                        else:               
                            worksheet.merge_range(f"{st_cell}{11}:{end_cell}{11}", item_case.flowrate, cell_format)
                            worksheet.merge_range(f"{st_cell}{12}:{end_cell}{12}", item_case.inletPressure, cell_format)
                            worksheet.merge_range(f"{st_cell}{13}:{end_cell}{13}", item_case.outletPressure, cell_format)
                            worksheet.merge_range(f"{st_cell}{14}:{end_cell}{14}", item_case.inletTemp, cell_format)
                            worksheet.merge_range(f"{st_cell}{15}:{end_cell}{15}", item_case.molecularWeight, cell_format)
                            worksheet.merge_range(f"{st_cell}{16}:{end_cell}{16}", item_case.specificHeatRatio, cell_format)
                            worksheet.merge_range(f"{st_cell}{17}:{end_cell}{17}", item_case.compressibility, cell_format)
                            worksheet.merge_range(f"{st_cell}{18}:{end_cell}{18}", item_case.xt, cell_format)
                            worksheet.merge_range(f"{st_cell}{19}:{end_cell}{19}", item_case.calculatedCv, cell_format)
                            worksheet.merge_range(f"{st_cell}{20}:{end_cell}{20}", item_case.openingPercentage, cell_format)
                            worksheet.merge_range(f"{st_cell}{21}:{end_cell}{21}", item_case.spl, cell_format)
                            worksheet.merge_range(f"{st_cell}{22}:{end_cell}{22}", item_case.x_delp, cell_format)
                            worksheet.merge_range(f"{st_cell}{23}:{end_cell}{23}", item_case.chokedDrop, cell_format)
                            worksheet.merge_range(f"{st_cell}{24}:{end_cell}{24}", item_case.y_expansion, cell_format)
                            worksheet.merge_range(f"{st_cell}{25}:{end_cell}{25}", item_case.xtp, cell_format)
                            worksheet.merge_range(f"{st_cell}{26}:{end_cell}{26}", item_case.fk, cell_format)
                            worksheet.merge_range(f"{st_cell}{27}:{end_cell}{27}", item_case.Fp, cell_format)
                            worksheet.merge_range(f"{st_cell}{28}:{end_cell}{28}", item_case.ar, cell_format)
                            worksheet.merge_range(f"{st_cell}{29}:{end_cell}{29}", item_case.reNumber, cell_format)
                            worksheet.merge_range(f"{st_cell}{30}:{end_cell}{30}", item_case.valveSize, cell_format)
                            worksheet.merge_range(f"{st_cell}{31}:{end_cell}{31}", item_case.pipeInVel, cell_format)
                            worksheet.merge_range(f"{st_cell}{32}:{end_cell}{32}", item_case.machNoUp, cell_format)
                            worksheet.merge_range(f"{st_cell}{33}:{end_cell}{33}", item_case.machNoDown, cell_format)
                            worksheet.merge_range(f"{st_cell}{34}:{end_cell}{34}", item_case.machNoValve, cell_format)
                            worksheet.merge_range(f"{st_cell}{35}:{end_cell}{35}", item_case.sonicVelUp, cell_format)
                            worksheet.merge_range(f"{st_cell}{36}:{end_cell}{36}", item_case.sonicVelDown, cell_format)
                            worksheet.merge_range(f"{st_cell}{37}:{end_cell}{37}", item_case.sonicVelValve, cell_format)
                            worksheet.merge_range(f"{st_cell}{38}:{end_cell}{38}", item_case.outletDensity, cell_format)
                            worksheet.merge_range(f"{st_cell}{39}:{end_cell}{39}", item_case.tex, cell_format)
                            worksheet.merge_range(f"{st_cell}{40}:{end_cell}{40}", item_case.powerLevel, cell_format)
                            worksheet.merge_range(f"{st_cell}{41}:{end_cell}{41}", item_case.requiredStages, cell_format)
                            worksheet.merge_range(f"{st_cell}{42}:{end_cell}{42}", '', cell_format)
       

                    else:
                        if i==2:
                            worksheet.write(f"{st_cell}{11}", '', cell_format)
                            worksheet.write(f"{st_cell}{12}", '', cell_format)
                            worksheet.write(f"{st_cell}{13}", '', cell_format)
                            worksheet.write(f"{st_cell}{14}", '',cell_format)
                            worksheet.write(f"{st_cell}{15}", '', cell_format)
                            worksheet.write(f"{st_cell}{16}", '',cell_format)
                            worksheet.write(f"{st_cell}{17}", '', cell_format)
                            worksheet.write(f"{st_cell}{18}", '',cell_format)
                            worksheet.write(f"{st_cell}{19}", '', cell_format)
                            worksheet.write(f"{st_cell}{20}", '', cell_format)
                            worksheet.write(f"{st_cell}{21}", '', cell_format)
                            worksheet.write(f"{st_cell}{22}", '',cell_format)
                            worksheet.write(f"{st_cell}{23}", '', cell_format)
                            worksheet.write(f"{st_cell}{24}", '', cell_format)
                            worksheet.write(f"{st_cell}{25}", '', cell_format)
                            worksheet.write(f"{st_cell}{26}", '', cell_format)
                            worksheet.write(f"{st_cell}{27}", '', cell_format)
                            worksheet.write(f"{st_cell}{28}", '', cell_format)
                            worksheet.write(f"{st_cell}{29}", '', cell_format)
                            worksheet.write(f"{st_cell}{30}", '', cell_format)
                            worksheet.write(f"{st_cell}{31}", '', cell_format)
                            worksheet.write(f"{st_cell}{32}", '', cell_format)
                            worksheet.write(f"{st_cell}{33}", '', cell_format)
                            worksheet.write(f"{st_cell}{34}", '', cell_format)
                            worksheet.write(f"{st_cell}{35}", '', cell_format)
                            worksheet.write(f"{st_cell}{36}", '', cell_format)
                            worksheet.write(f"{st_cell}{37}", '', cell_format)
                            worksheet.write(f"{st_cell}{38}", '', cell_format)
                            worksheet.write(f"{st_cell}{39}", '', cell_format)
                            worksheet.write(f"{st_cell}{40}", '', cell_format)
                            worksheet.write(f"{st_cell}{41}", '', cell_format)
                            worksheet.write(f"{st_cell}{42}", '', cell_format)
                            # worksheet.write(f"{st_cell}{43}", '', cell_format)
                            # worksheet.write(f"{st_cell}{44}", '', cell_format)
                            # worksheet.write(f"{st_cell}{45}", '', cell_format)
                            # worksheet.write(f"{st_cell}{46}", '', cell_format)
                            # worksheet.write(f"{st_cell}{47}", '', cell_format)
                        else:
                            worksheet.merge_range(f"{st_cell}{11}:{end_cell}{11}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{12}:{end_cell}{12}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{13}:{end_cell}{13}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{14}:{end_cell}{14}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{15}:{end_cell}{15}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{16}:{end_cell}{16}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{17}:{end_cell}{17}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{18}:{end_cell}{18}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{19}:{end_cell}{19}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{20}:{end_cell}{20}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{21}:{end_cell}{21}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{22}:{end_cell}{22}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{23}:{end_cell}{23}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{24}:{end_cell}{24}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{25}:{end_cell}{25}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{26}:{end_cell}{26}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{27}:{end_cell}{27}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{28}:{end_cell}{28}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{29}:{end_cell}{29}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{30}:{end_cell}{30}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{31}:{end_cell}{31}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{32}:{end_cell}{32}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{33}:{end_cell}{33}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{34}:{end_cell}{34}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{35}:{end_cell}{35}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{36}:{end_cell}{36}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{37}:{end_cell}{37}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{38}:{end_cell}{38}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{39}:{end_cell}{39}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{40}:{end_cell}{40}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{41}:{end_cell}{41}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{42}:{end_cell}{42}", '', cell_format)
                            # worksheet.merge_range(f"{st_cell}{43}:{end_cell}{43}", '', cell_format)
                            # worksheet.merge_range(f"{st_cell}{44}:{end_cell}{44}", '', cell_format)
                            # worksheet.merge_range(f"{st_cell}{45}:{end_cell}{45}", '', cell_format)
                            # worksheet.merge_range(f"{st_cell}{46}:{end_cell}{46}", '', cell_format)
                            # worksheet.merge_range(f"{st_cell}{47}:{end_cell}{47}", '', cell_format)
            
            
            elif fluid_type == 'Liquid':
                case_cv_values = 0
                cv_graph = 0; cv_graph_values = [];case_cv_final =[];case_cv_finalPercent = []
                for i in range(5):
                    st_cell,end_cell = case_cell[i].split(":") 
                    
                    if i < len(itemCase): 

                        item_case = itemCase[i] 
                        case_cv_final.append(item_case.calculatedCv)
                        case_cv_finalPercent.append(item_case.openingPercentage)
                        if item_case.valveDiaId:
                            # case_cv_values
                            cv_graph = 1

                            cv_element = getDBElementWithId(cvTable, item_case.valveDiaId)
                            # db.session.query(cvValues).filter_by(cv=cv_element, coeff='Cv').first()
                            item_case_cvs = db.session.query(cvValues).filter_by(cv=cv_element, coeff='Cv').first()
                            print(f'itemcasecvssssddvv {item_case_cvs}')
                            cv_graph_values = [item_case_cvs.one, item_case_cvs.two, item_case_cvs.three,item_case_cvs.four,item_case_cvs.five,item_case_cvs.six,item_case_cvs.seven,item_case_cvs.eight,item_case_cvs.nine,item_case_cvs.ten]

                        if i == 2:
                            worksheet.write(f"{st_cell}{11}", item_case.flowrate, cell_format)
                            worksheet.write(f"{st_cell}{12}", item_case.inletPressure, cell_format)
                            worksheet.write(f"{st_cell}{13}:", item_case.outletPressure, cell_format)
                            worksheet.write(f"{st_cell}{14}:", item_case.inletTemp, cell_format)
                            worksheet.write(f"{st_cell}{15}:", item_case.specificGravity, cell_format)
                            worksheet.write(f"{st_cell}{16}:", item_case.kinematicViscosity, cell_format)
                            worksheet.write(f"{st_cell}{17}:", item_case.vaporPressure, cell_format)
                            worksheet.write(f"{st_cell}{18}:", item_case.fl, cell_format)
                            worksheet.write(f"{st_cell}{19}:", item_case.calculatedCv, cell_format)
                            worksheet.write(f"{st_cell}{20}:", item_case.openingPercentage, cell_format)
                            worksheet.write(f"{st_cell}{21}:", item_case.spl, cell_format)
                            worksheet.write(f"{st_cell}{22}:", item_case.x_delp, cell_format)
                            worksheet.write(f"{st_cell}{23}:", item_case.chokedDrop, cell_format)
                            worksheet.write(f"{st_cell}{24}:", item_case.Ff, cell_format)
                            worksheet.write(f"{st_cell}{25}:", item_case.Fp, cell_format)
                            worksheet.write(f"{st_cell}{26}:", item_case.Flp, cell_format)
                            worksheet.write(f"{st_cell}{27}:", item_case.kc, cell_format)
                            worksheet.write(f"{st_cell}{28}:", item_case.ar, cell_format)
                            worksheet.write(f"{st_cell}{29}:", item_case.reNumber, cell_format)
                            worksheet.write(f"{st_cell}{30}:", item_case.valveSize, cell_format)
                            worksheet.write(f"{st_cell}{31}:", item_case.pipeInVel, cell_format)
                            worksheet.write(f"{st_cell}{32}:", item_case.pipeOutVel, cell_format)
                            worksheet.write(f"{st_cell}{33}:", item_case.valveVel, cell_format)
                            worksheet.write(f"{st_cell}{34}:", item_case.tex, cell_format)
                            worksheet.write(f"{st_cell}{35}:", item_case.powerLevel, cell_format)
                            worksheet.write(f"{st_cell}{36}:", item_case.requiredStages, cell_format)
                            worksheet.write(f"{st_cell}{37}:", '', cell_format)           

                        else:               
                            worksheet.merge_range(f"{st_cell}{11}:{end_cell}{11}", item_case.flowrate, cell_format)
                            worksheet.merge_range(f"{st_cell}{12}:{end_cell}{12}", item_case.inletPressure, cell_format)
                            worksheet.merge_range(f"{st_cell}{13}:{end_cell}{13}", item_case.outletPressure, cell_format)
                            worksheet.merge_range(f"{st_cell}{14}:{end_cell}{14}", item_case.inletTemp, cell_format)
                            worksheet.merge_range(f"{st_cell}{15}:{end_cell}{15}", item_case.specificGravity, cell_format)
                            worksheet.merge_range(f"{st_cell}{16}:{end_cell}{16}", item_case.kinematicViscosity, cell_format)
                            worksheet.merge_range(f"{st_cell}{17}:{end_cell}{17}", item_case.vaporPressure, cell_format)
                            worksheet.merge_range(f"{st_cell}{18}:{end_cell}{18}", item_case.fl, cell_format)
                            worksheet.merge_range(f"{st_cell}{19}:{end_cell}{19}", item_case.calculatedCv, cell_format)
                            worksheet.merge_range(f"{st_cell}{20}:{end_cell}{20}", item_case.openingPercentage, cell_format)
                            worksheet.merge_range(f"{st_cell}{21}:{end_cell}{21}", item_case.spl, cell_format)
                            worksheet.merge_range(f"{st_cell}{22}:{end_cell}{22}", item_case.x_delp, cell_format)
                            worksheet.merge_range(f"{st_cell}{23}:{end_cell}{23}", item_case.chokedDrop, cell_format)
                            worksheet.merge_range(f"{st_cell}{24}:{end_cell}{24}", item_case.Ff, cell_format)
                            worksheet.merge_range(f"{st_cell}{25}:{end_cell}{25}", item_case.Fp, cell_format)
                            worksheet.merge_range(f"{st_cell}{26}:{end_cell}{26}", item_case.Flp, cell_format)
                            worksheet.merge_range(f"{st_cell}{27}:{end_cell}{27}", item_case.kc, cell_format)
                            worksheet.merge_range(f"{st_cell}{28}:{end_cell}{28}", item_case.ar, cell_format)
                            worksheet.merge_range(f"{st_cell}{29}:{end_cell}{29}", item_case.reNumber, cell_format)
                            worksheet.merge_range(f"{st_cell}{30}:{end_cell}{30}", item_case.valveSize, cell_format)
                            worksheet.merge_range(f"{st_cell}{31}:{end_cell}{31}", item_case.pipeInVel, cell_format)
                            worksheet.merge_range(f"{st_cell}{32}:{end_cell}{32}", item_case.pipeOutVel, cell_format)
                            worksheet.merge_range(f"{st_cell}{33}:{end_cell}{33}", item_case.valveVel, cell_format)
                            worksheet.merge_range(f"{st_cell}{34}:{end_cell}{34}", item_case.tex, cell_format)
                            worksheet.merge_range(f"{st_cell}{35}:{end_cell}{35}", item_case.powerLevel, cell_format)
                            worksheet.merge_range(f"{st_cell}{36}:{end_cell}{36}", item_case.requiredStages, cell_format)
                            worksheet.merge_range(f"{st_cell}{37}:{end_cell}{37}", '', cell_format)           

                    else:
                        if i==2:
                            worksheet.write(f"{st_cell}{11}", '', cell_format)
                            worksheet.write(f"{st_cell}{12}", '', cell_format)
                            worksheet.write(f"{st_cell}{13}", '', cell_format)
                            worksheet.write(f"{st_cell}{14}", '', cell_format)
                            worksheet.write(f"{st_cell}{15}", '', cell_format)
                            worksheet.write(f"{st_cell}{16}", '', cell_format)
                            worksheet.write(f"{st_cell}{17}", '', cell_format)
                            worksheet.write(f"{st_cell}{18}", '', cell_format)
                            worksheet.write(f"{st_cell}{19}", '', cell_format)
                            worksheet.write(f"{st_cell}{20}", '', cell_format)
                            worksheet.write(f"{st_cell}{21}", '', cell_format)
                            worksheet.write(f"{st_cell}{22}", '', cell_format)
                            worksheet.write(f"{st_cell}{23}", '', cell_format)
                            worksheet.write(f"{st_cell}{24}", '', cell_format)
                            worksheet.write(f"{st_cell}{25}", '', cell_format)
                            worksheet.write(f"{st_cell}{26}", '', cell_format)
                            worksheet.write(f"{st_cell}{27}", '', cell_format)
                            worksheet.write(f"{st_cell}{28}", '', cell_format)
                            worksheet.write(f"{st_cell}{29}", '', cell_format)
                            worksheet.write(f"{st_cell}{30}", '', cell_format)
                            worksheet.write(f"{st_cell}{31}", '', cell_format)
                            worksheet.write(f"{st_cell}{32}", '', cell_format)
                            worksheet.write(f"{st_cell}{33}", '', cell_format)
                            worksheet.write(f"{st_cell}{34}", '', cell_format)
                            worksheet.write(f"{st_cell}{35}", '', cell_format)
                            worksheet.write(f"{st_cell}{36}", '', cell_format)
                            worksheet.write(f"{st_cell}{37}", '', cell_format)
                        
                        else:
                            worksheet.merge_range(f"{st_cell}{11}:{end_cell}{11}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{12}:{end_cell}{12}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{13}:{end_cell}{13}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{14}:{end_cell}{14}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{15}:{end_cell}{15}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{16}:{end_cell}{16}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{17}:{end_cell}{17}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{18}:{end_cell}{18}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{19}:{end_cell}{19}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{20}:{end_cell}{20}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{21}:{end_cell}{21}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{22}:{end_cell}{22}", '',cell_format)
                            worksheet.merge_range(f"{st_cell}{23}:{end_cell}{23}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{24}:{end_cell}{24}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{25}:{end_cell}{25}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{26}:{end_cell}{26}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{27}:{end_cell}{27}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{28}:{end_cell}{28}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{29}:{end_cell}{29}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{30}:{end_cell}{30}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{31}:{end_cell}{31}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{32}:{end_cell}{32}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{33}:{end_cell}{33}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{34}:{end_cell}{34}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{35}:{end_cell}{35}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{36}:{end_cell}{36}", '', cell_format)
                            worksheet.merge_range(f"{st_cell}{37}:{end_cell}{37}", '', cell_format)
                
                print(f'cvGraphVakuesffffffddd {cv_graph_values}')
                print(f'FINAL CV CALC {case_cv_final}')

                
            print(f'CVGRAPHVALUES {cv_graph_values}')
            if fluid_type == 'Gas' and cv_graph_values:
                # worksheet.write("F44", 'Cv')
                print('header___sss',header_details)
                worksheet.merge_range("D61:E61", 'Valve Style', cell_format)
                worksheet.merge_range("D62:E62", 'Trim Type', cell_format)
                worksheet.merge_range("D63:E63", 'Trim Characteristics', cell_format)

                worksheet.merge_range("F61:H61",  header_details[f_cnt][14], cell_format)
                worksheet.merge_range("F62:H62", header_details[f_cnt][15], cell_format)
                worksheet.merge_range("F63:H63", header_details[f_cnt][16], cell_format)

                def getCVplotchart(opencv_final, cv_values, calc_opencv, calc_cvvalues):
                    cnt = 0
                    for i in calc_opencv:
                        for j in range(len(opencv_final)):
                            if opencv_final[j] > i:
                                opencv_final.insert(j,i)
                                cv_values.insert(j,calc_cvvalues[cnt])
                                cnt+=1
                                break
                    calc_finalcv = []
                    for i in cv_values:
                        if i in calc_cvvalues:
                            calc_finalcv.append(i)
                        else:
                            calc_finalcv.append(None)

         
                    print(f'openingpercentagedatas {opencv_final},{cv_values},{calc_finalcv}')
                    final_xy = {
                        'x_values' : opencv_final,
                        'y_values' : cv_values,
                        'y1_values' : calc_finalcv
                    }
                    return final_xy


                cv_values = cv_graph_values
                open_cv = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
                calc_opencv = case_cv_finalPercent
                calc_cvvalues = case_cv_final
                sheet_number = f_cnt+1

                final_values = getCVplotchart(open_cv,cv_values,calc_opencv,calc_cvvalues)
                print(f'LLSJSJJSJSJJ {final_values}')
             
                # Write the data to the worksheet
                worksheet.write_column('C45', final_values['y_values'])
                worksheet.write_column('D45', final_values['x_values'])
                worksheet.write_column('E45', final_values['y1_values'])    


                # Add a line chart
                scatter_chart = workbook.add_chart({"type": "scatter", "name": "Scatter Chart", "embedded": True})

                # Add the scatter series to the chart
                scatter_series = {
                    'categories': f'=Sheet{sheet_number}!$D$45:$D$58',  # X-axis data for scatter plot
                    'values': f'=Sheet{sheet_number}!$E$45:$E$58',       # Y-axis data for scatter plot
                    'marker': {'type': 'circle', 'size': 5, 'fill': {'color': '#FF0000'}}, 
                    'line': {'none': True},
                    'name': 'Cases CV'
                    
                    # Marker style for scatter plot
                }
                scatter_chart.add_series(scatter_series)



                # Add a line series to the scatter chart
                line_series = {
                    'categories': f'=Sheet{sheet_number}!$D$45:$D$58',  # X-axis data for line plot
                    'values': f'=Sheet{sheet_number}!$C$45:$C$58',
                    'marker': {'type': 'none'},
                    'line': {'width': 1.5 ,'color':'#0070C0'},     # Y-axis data for line plot
                    'name': 'Std CV'
                }
                scatter_chart.add_series(line_series)

                scatter_chart.set_y_axis({
                    'min': 0 ,           # Minimum X-axis value
                    'max': cv_graph_values[-1],          # Maximum X-axis value
                    'major_unit': cv_graph_values[-1] / 10,        # Interval between major ticks
                    'minor_unit': 1,     # Interval between minor ticks
                    'minor_unit_type': 'num',  # Set minor unit type to number
                    'num_font': {'size': 10},  # Font size for tick labels
                    'name': 'Flow Coefficient, Cv',  # Name of the X-axis
                    'major_gridlines': {'visible': True, 'line': {'color': '#d1d1d1', 'extend': False} },
                    'num_font':  {'name': 'Arial', 'size': 9, 'color': '#6c757d'},
                    'name_font' : {'name': 'Arial', 'size': 10, 'color': '#6c757d','bold':False},
                    'crossing': 0, 
                })

                
         

                # Set x-axis range and tick intervals for the combined chart
                scatter_chart.set_x_axis({
                    'min': 0,           # Minimum X-axis value
                    'max': 100,          # Maximum X-axis value
                    'major_unit': 10,    # Interval between major ticks
                    'minor_unit': 1,     # Interval between minor ticks
                    'minor_unit_type': 'num',  # Set minor unit type to number  
                    'name': '% Opening',  # Name of the X-axis
                    'major_gridlines': {'visible': True, 'line': {'color': '#d1d1d1', 'extend': False} },
                    'num_font':  {'name': 'Arial', 'size': 9, 'color': '#6c757d'},
                    'name_font' : {'name': 'Arial', 'size': 10, 'color': '#6c757d', 'bold':False},
                    'crossing': 0, 
                    
                })



               

                worksheet.insert_chart('C45', scatter_chart)
                worksheet.write('A71:B71', 'FLOW CONTROL COMMUNE',h3)
                worksheet.write('M71', 'FR/AE/004/A Iss.01', f1)


            
            elif fluid_type == 'Liquid' and cv_graph_values:
                print('header___',header_details)
                worksheet.merge_range("D55:E55", 'Valve Style', cell_format)
                worksheet.merge_range("D56:E56", 'Trim Type', cell_format)
                worksheet.merge_range("D57:E57", 'Trim Characteristics', cell_format)

                worksheet.merge_range("F55:H55",  header_details[f_cnt][14], cell_format)
                worksheet.merge_range("F56:H56", header_details[f_cnt][15], cell_format)
                worksheet.merge_range("F57:H57", header_details[f_cnt][16], cell_format)
                def getCVplotchart(opencv_final, cv_values, calc_opencv, calc_cvvalues):
                    cnt = 0
                    for i in calc_opencv:
                        for j in range(len(opencv_final)):
                            if opencv_final[j] > i:
                                opencv_final.insert(j, i)
                                cv_values.insert(j, calc_cvvalues[cnt])
                                cnt += 1
                                break
                    calc_finalcv = []
                    for i in cv_values:
                        if i in calc_cvvalues:
                            calc_finalcv.append(i)
                        else:
                            calc_finalcv.append(None)

                    print(f'openingpercentagedatas {opencv_final},{cv_values},{calc_finalcv}')
                    final_xy = {
                        'x_values': opencv_final,
                        'y_values': cv_values,
                        'y1_values': calc_finalcv
                    }
                    return final_xy

                print(f'cv_grphvalues {cv_graph_values}')
                cv_values = cv_graph_values
                open_cv = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
                calc_opencv = case_cv_finalPercent
                calc_cvvalues = case_cv_final
                sheet_number = f_cnt + 1

                final_values = getCVplotchart(open_cv, cv_values, calc_opencv, calc_cvvalues)
                print(f'LLSJSJJSJSJJ {final_values}')

                # Write the data to the worksheet
                worksheet.write_column('C40', final_values['y_values'])
                worksheet.write_column('D40', final_values['x_values'])
                worksheet.write_column('E40', final_values['y1_values'])

                # Add a line chart
                scatter_chart = workbook.add_chart({"type": "scatter", "name": "Scatter Chart", "subtype": "straight", "embedded": True})

                # Add the scatter series to the chart
                scatter_series = {
                    'categories': f'=Sheet{sheet_number}!$D$40:$D$54',  # X-axis data for scatter plot
                    'values': f'=Sheet{sheet_number}!$E$40:$E$54',       # Y-axis data for scatter plot
                    'marker': {'type': 'circle', 'size': 5, 'fill': {'color': '#FF0000'}},
                    'line': {'none': True},
                    'name': 'Cases CV'
                    # Marker style for scatter plot
                }
                scatter_chart.add_series(scatter_series)

                # Add a line series to the scatter chart
                line_series = {
                    'categories': f'=Sheet{sheet_number}!$D$40:$D$54',  # X-axis data for line plot
                    'values': f'=Sheet{sheet_number}!$C$40:$C$54',
                    'marker': {'type': 'none'},
                    'line': {'width': 1.5, 'color': '#0070C0'},     # Y-axis data for line plot
                    'name': 'Std CV'
                }
                scatter_chart.add_series(line_series)

                scatter_chart.set_y_axis({
                    'min': 0,           # Minimum X-axis value
                    'max': cv_graph_values[-1] ,          # Maximum X-axis value
                    'major_unit': cv_graph_values[-1] / 10,        # Interval between major ticks
                    'minor_unit': 1,     # Interval between minor ticks
                    'minor_unit_type': 'num',  # Set minor unit type to number
                    'num_font': {'size': 10},  # Font size for tick labels
                    'name': 'Percentage Openings / Degree',  # Name of the X-axis
                    'major_gridlines': {'visible': True, 'line': {'color': '#d1d1d1', 'extend': False}},
                    'num_font': {'name': 'Arial', 'size': 9, 'color': '#6c757d'},
                    'name_font': {'name': 'Arial', 'size': 10, 'color': '#6c757d', 'bold': False},
                    'crossing': 0, 
                })

                # Set x-axis range and tick intervals for the combined chart
                scatter_chart.set_x_axis({
                    'min': 0,           # Minimum X-axis value (adjusted from 10 to 0)
                    'max': 100,         # Maximum X-axis value
                    'major_unit': 10,   # Interval between major ticks
                    'minor_unit': 1,    # Interval between minor ticks
                    'minor_unit_type': 'num',  # Set minor unit type to number
                    'num_font': {'size': 10},  # Font size for tick labels
                    'name': 'Percentage Openings / Degree',  # Name of the X-axis
                    'major_gridlines': {'visible': True, 'line': {'color': '#d1d1d1', 'extend': False}},
                    'num_font': {'name': 'Arial', 'size': 9, 'color': '#6c757d', 'line': {'extend': False}},
                    'name_font': {'name': 'Arial', 'size': 10, 'color': '#6c757d', 'bold': False},
                    'crossing': 0, 
                })

                worksheet.insert_chart('C40', scatter_chart)

                worksheet.write('A65:B65', 'FLOW CONTROL COMMUNE PVT LTD', h3)
                worksheet.write('M65', 'FR/AE/004/A Iss.01', f1)
                # worksheet.write('L72', 'FR/AE/004/B Iss.01', f1)
                # Continue for other fields
            # if item_case1:
            #     worksheet.merge_range("E11:F11", item_case1.flowrate, cell_format)
            #     worksheet.merge_range("E12:F12", item_case1.inletPressure, cell_format)
            #     worksheet.merge_range("E13:F13", item_case1.outletPressure, cell_format)
            #     worksheet.merge_range("E14:F14", item_case1.inletTemp, cell_format)
            #     worksheet.merge_range("E15:F15", item_case1.molecularWeight, cell_format)
            #     worksheet.merge_range("E16:F16", item_case1.specificHeatRatio, cell_format)
            #     worksheet.merge_range("E17:F17", item_case1.compressibility, cell_format)
            #     worksheet.merge_range("E18:F18", item_case1.xt, cell_format)
            #     worksheet.merge_range("E19:F19", item_case1.calculatedCv, cell_format)
            #     worksheet.merge_range("E20:F20", item_case1.openingPercentage, cell_format)
            #     worksheet.merge_range("E21:F21", '',cell_format)
            #     worksheet.merge_range("E22:F22", item_case1.x_delp,cell_format)
            #     worksheet.merge_range("E23:F23", item_case1.chokedDrop, cell_format)
            #     worksheet.merge_range("E24:F24", item_case1.y_expansion, cell_format)
            #     worksheet.merge_range("E25:F25", item_case1.xtp, cell_format)
            #     worksheet.merge_range("E26:F26", item_case1.fk, cell_format)
            #     worksheet.merge_range("E27:F27", item_case1.Ff, cell_format)
            #     worksheet.merge_range("E28:F28", item_case1.Fp, cell_format)
            #     worksheet.merge_range("E29:F29", item_case1.Flp, cell_format)
            #     worksheet.merge_range("E30:F30", item_case1.kc, cell_format)
            #     worksheet.merge_range("E31:F31", item_case1.ar, cell_format)
            #     worksheet.merge_range("E32:F32", item_case1.reNumber, cell_format)
            #     worksheet.merge_range("E33:F33", '', cell_format)
            #     worksheet.merge_range("E34:F34", '', cell_format)
            #     worksheet.merge_range("E35:F35", item_case1.pipeInVel, cell_format)
            #     worksheet.merge_range("E36:F36", item_case1.pipeOutVel, cell_format)
            #     worksheet.merge_range("E37:F37", '', cell_format)
            #     worksheet.merge_range("E38:F38", item_case1.machNoUp, cell_format)
            #     worksheet.merge_range("E39:F39", item_case1.machNoDown, cell_format)
            #     worksheet.merge_range("E40:F40", item_case1.machNoValve, cell_format)
            #     worksheet.merge_range("E41:F41", item_case1.sonicVelUp, cell_format)
            #     worksheet.merge_range("E42:F42", item_case1.sonicVelDown, cell_format)
            #     worksheet.merge_range("E43:F43", item_case1.sonicVelValve, cell_format)
            #     worksheet.merge_range("E44:F44", item_case1.outletDensity, cell_format)
            #     worksheet.merge_range("E45:F45", '', cell_format)
            #     worksheet.merge_range("E46:F46", '', cell_format)
            #     worksheet.merge_range("E47:F47", '', cell_format)



            worksheet.write('G1', '', )
            worksheet.write('G2', '', )
            worksheet.write('G8', '', )
            worksheet.merge_range("G9:H9", 'Critical Pressure.', cell_format)
            worksheet.merge_range("G10:H10", 'Case 2', bold)
            # worksheet.merge_range("G11:H11", '', cell_format)
            # worksheet.merge_range("G12:H12", '', cell_format)
            # worksheet.merge_range("G13:H13", '', cell_format)
            # worksheet.merge_range("G14:H14", '', cell_format)
            # worksheet.merge_range("G15:H15", '', cell_format)
            # worksheet.merge_range("G16:H16", '', cell_format)
            # worksheet.merge_range("G17:H17", '', cell_format)
            # worksheet.merge_range("G18:H18", '', cell_format)
            # worksheet.merge_range("G19:H19", '', cell_format)
            # worksheet.merge_range("G20:H20", '', cell_format)
            # worksheet.merge_range("G21:H21", '', cell_format)
            # worksheet.merge_range("G22:H22", '', cell_format)
            # worksheet.merge_range("G23:H23", '', cell_format)
            # worksheet.merge_range("G24:H24", '', cell_format)
            # worksheet.merge_range("G25:H25", '', cell_format)
            # worksheet.merge_range("G26:H26", '', cell_format)
            # worksheet.merge_range("G27:H27", '', cell_format)
            # worksheet.merge_range("G28:H28", '', cell_format)
            # worksheet.merge_range("G29:H29", '', cell_format)
            # worksheet.merge_range("G30:H30", '', cell_format)
            # worksheet.merge_range("G31:H31", '', cell_format)
            # worksheet.merge_range("G32:H32", '', cell_format)
            # worksheet.merge_range("G33:H33", '', cell_format)
            # worksheet.merge_range("G34:H34", '', cell_format)
            # worksheet.merge_range("G35:H35", '', cell_format)
            # worksheet.merge_range("G36:H36", '', cell_format)
            # worksheet.merge_range("G37:H37", '', cell_format)
            # worksheet.merge_range("G38:H38", '', cell_format)
            # worksheet.merge_range("G39:H39", '', cell_format)
            # worksheet.merge_range("G40:H40", '', cell_format)
            # worksheet.merge_range("G41:H41", '', cell_format)
            # worksheet.merge_range("G42:H42", '', cell_format)
            # worksheet.merge_range("G43:H43", '', cell_format)
            # worksheet.merge_range("G44:H44", '', cell_format)
            # worksheet.merge_range("G45:H45", '', cell_format)
            # worksheet.merge_range("G46:H46", '', cell_format)
            # worksheet.merge_range("G47:H47", '', cell_format)


            worksheet.write('I1', '')
            worksheet.write('I2', '')
            worksheet.write('I3', 'Quote No', cell_format)
            worksheet.write('I4', 'W/O No', cell_format)
            worksheet.write('I5', 'Serial No', cell_format)
            worksheet.write('I6', 'Tag No', cell_format)
            worksheet.write('I7', 'Item No / Qty', cell_format)
            # worksheet.write('I8', '', cell_format)
            worksheet.write('I9', '', cell_format)
            worksheet.write('I10', 'Case 3', bold)
            # worksheet.write('I11', '', cell_format)
            # worksheet.write('I12', '', cell_format)
            # worksheet.write('I13', '', cell_format)
            # worksheet.write('I14', '', cell_format)
            # worksheet.write('I15', '', cell_format)
            # worksheet.write('I16', '', cell_format)
            # worksheet.write('I17', '', cell_format)
            # worksheet.write('I18', '', cell_format)
            # worksheet.write('I19', '', cell_format)
            # worksheet.write('I20', '', cell_format)
            # worksheet.write('I21', '', cell_format)
            # worksheet.write('I22', '', cell_format)
            # worksheet.write('I23', '', cell_format)
            # worksheet.write('I24', '', cell_format)
            # worksheet.write('I25', '', cell_format)
            # worksheet.write('I26', '', cell_format)
            # worksheet.write('I27', '', cell_format)
            # worksheet.write('I28', '', cell_format)
            # worksheet.write('I29', '', cell_format)
            # worksheet.write('I30', '', cell_format)
            # worksheet.write('I31', '', cell_format)
            # worksheet.write('I32', '', cell_format)
            # worksheet.write('I33', '', cell_format)
            # worksheet.write('I34', '', cell_format)
            # worksheet.write('I35', '', cell_format)
            # worksheet.write('I36', '', cell_format)
            # worksheet.write('I37', '', cell_format)
            # worksheet.write('I38', '', cell_format)
            # worksheet.write('I39', '', cell_format)
            # worksheet.write('I40', '', cell_format)
            # worksheet.write('I41', '', cell_format)
            # worksheet.write('I42', '', cell_format)
            # worksheet.write('I43', '', cell_format)
            # worksheet.write('I44', '', cell_format)
            # worksheet.write('I45', '', cell_format)
            # worksheet.write('I46', '', cell_format)
            # worksheet.write('I47', '', cell_format)


            worksheet.merge_range("J3:M3", '', cell_format)
            worksheet.merge_range("J4:M4", '', cell_format)
            worksheet.merge_range("J5:M5", '', cell_format)
            worksheet.merge_range("J6:M6", '', cell_format)
            worksheet.merge_range("J7:M7", '', cell_format)
            worksheet.write('J3', header_details[f_cnt][5], cell_format1)
            worksheet.write('J4', header_details[f_cnt][6], cell_format1)
            worksheet.write('J5', header_details[f_cnt][7], cell_format1)
            worksheet.write('J6', header_details[f_cnt][8], cell_format1)
            worksheet.write('J7', header_details[f_cnt][9], cell_format1)

            
            worksheet.merge_range("L9:M9", '', cell_format)
            worksheet.merge_range("L10:M10", 'Case 5', bold)
            

            worksheet.write('C8', header_details[f_cnt][10], cell_format1)
            worksheet.write('C9', header_details[f_cnt][11], cell_format1)
            worksheet.write('I9', header_details[f_cnt][12], cell_format1)
            worksheet.write('L9', header_details[f_cnt][13], cell_format1)
           
            # worksheet.merge_range("J7:M7", '', cell_format)
            worksheet.merge_range("J9:K9", 'Shutoff Pressure.', cell_format)
            worksheet.merge_range("J10:K10", 'Case 4', bold)
            # worksheet.merge_range("J11:K11", '', cell_format)
            # worksheet.merge_range("J12:K12", '', cell_format)
            # worksheet.merge_range("J13:K13", '', cell_format)
            # worksheet.merge_range("J14:K14", '', cell_format)
            # worksheet.merge_range("J15:K15", '', cell_format)
            # worksheet.merge_range("J16:K16", '', cell_format)
            # worksheet.merge_range("J17:K17", '', cell_format)
            # worksheet.merge_range("J18:K18", '', cell_format)
            # worksheet.merge_range("J19:K19", '', cell_format)
            # worksheet.merge_range("J20:K20", '', cell_format)
            # worksheet.merge_range("J21:K21", '', cell_format)
            # worksheet.merge_range("J22:K22", '', cell_format)
            # worksheet.merge_range("J23:K23", '', cell_format)
            # worksheet.merge_range("J24:K24", '', cell_format)
            # worksheet.merge_range("J25:K25", '', cell_format)
            # worksheet.merge_range("J26:K26", '', cell_format)
            # worksheet.merge_range("J27:K27", '', cell_format)
            # worksheet.merge_range("J28:K28", '', cell_format)
            # worksheet.merge_range("J29:K29", '', cell_format)
            # worksheet.merge_range("J30:K30", '', cell_format)
            # worksheet.merge_range("J31:K31", '', cell_format)
            # worksheet.merge_range("J32:K32", '', cell_format)
            # worksheet.merge_range("J33:K33", '', cell_format)
            # worksheet.merge_range("J34:K34", '', cell_format)
            # worksheet.merge_range("J35:K35", '', cell_format)
            # worksheet.merge_range("J36:K36", '', cell_format)
            # worksheet.merge_range("J37:K37", '', cell_format)
            # worksheet.merge_range("J38:K38", '', cell_format)
            # worksheet.merge_range("J39:K39", '', cell_format)
            # worksheet.merge_range("J40:K40", '', cell_format)
            # worksheet.merge_range("J41:K41", '', cell_format)
            # worksheet.merge_range("J42:K42", '', cell_format)
            # worksheet.merge_range("J43:K43", '', cell_format)
            # worksheet.merge_range("J44:K44", '', cell_format)
            # worksheet.merge_range("J45:K45", '', cell_format)
            # worksheet.merge_range("J46:K46", '', cell_format)
            # worksheet.merge_range("J47:K47", '', cell_format)




            # worksheet.merge_range("L11:M11", '', cell_format)
            # worksheet.merge_range("L12:M12", '', cell_format)
            # worksheet.merge_range("L13:M13", '', cell_format)
            # worksheet.merge_range("L14:M14", '', cell_format)
            # worksheet.merge_range("L15:M15", '', cell_format)
            # worksheet.merge_range("L16:M16", '', cell_format)
            # worksheet.merge_range("L17:M17", '', cell_format)
            # worksheet.merge_range("L18:M18", '', cell_format)
            # worksheet.merge_range("L19:M19", '', cell_format)
            # worksheet.merge_range("L20:M20", '', cell_format)
            # worksheet.merge_range("L21:M21", '', cell_format)
            # worksheet.merge_range("L22:M22", '', cell_format)
            # worksheet.merge_range("L23:M23", '', cell_format)
            # worksheet.merge_range("L24:M24", '', cell_format)
            # worksheet.merge_range("L25:M25", '', cell_format)
            # worksheet.merge_range("L26:M26", '', cell_format)
            # worksheet.merge_range("L27:M27", '', cell_format)
            # worksheet.merge_range("L28:M28", '', cell_format)
            # worksheet.merge_range("L29:M29", '', cell_format)
            # worksheet.merge_range("L30:M30", '', cell_format)
            # worksheet.merge_range("L31:M31", '', cell_format)
            # worksheet.merge_range("L32:M32", '', cell_format)
            # worksheet.merge_range("L33:M33", '', cell_format)
            # worksheet.merge_range("L34:M34", '', cell_format)
            # worksheet.merge_range("L35:M35", '', cell_format)
            # worksheet.merge_range("L36:M36", '', cell_format)
            # worksheet.merge_range("L37:M37", '', cell_format)
            # worksheet.merge_range("L38:M38", '', cell_format)
            # worksheet.merge_range("L39:M39", '', cell_format)
            # worksheet.merge_range("L40:M40", '', cell_format)
            # worksheet.merge_range("L41:M41", '', cell_format)
            # worksheet.merge_range("L42:M42", '', cell_format)
            # worksheet.merge_range("L43:M43", '', cell_format)
            # worksheet.merge_range("L44:M44", '', cell_format)
            # worksheet.merge_range("L45:M45", '', cell_format)
            # worksheet.merge_range("L46:M46", '', cell_format)
            # worksheet.merge_range("L47:M47", '', cell_format)


            if fluid_type == 'Gas':
                worksheet.write("A43", 'BB',boldc)
                worksheet.write('A65', 'CC', boldc)
                worksheet.write('A66', '33', cell_format2)
                worksheet.write('A67', '34', cell_format2)
                worksheet.write('A68', '35', cell_format2)
                worksheet.write('A69', '36', cell_format2)
                worksheet.write('A70', '37', cell_format2)


                worksheet.merge_range("B43:M43", 'CV Plot', bold)
                worksheet.merge_range("B65:J65", 'Notes', bold)
                worksheet.merge_range("B66:J66", '', cell_format)
                worksheet.merge_range("B67:J67", '', cell_format)
                worksheet.merge_range("B68:J68", '', cell_format)   
                worksheet.merge_range("B69:J69", '', cell_format)
                worksheet.merge_range("B70:J70", '', cell_format)


                worksheet.merge_range("K65:M65", 'Revision Control', bold)
                worksheet.write('K66', 'Rev.', cell_format)
                worksheet.write('L66', 'By.', cell_format)
                worksheet.write('M66', 'Date.', cell_format)

                worksheet.write('K67', '', cell_format2)
                worksheet.write('K68', '', cell_format2)
                worksheet.write('K69', '', cell_format2)
                worksheet.write('K70', '', cell_format2)

                worksheet.write('L67', '', cell_format2)
                worksheet.write('L68', '', cell_format2)
                worksheet.write('L69', '', cell_format2)
                worksheet.write('L70', '', cell_format2)

                worksheet.write('M67', '', cell_format2)
                worksheet.write('M68', '', cell_format2)
                worksheet.write('M69', '', cell_format2)
                worksheet.write('M70', '', cell_format2)

            elif fluid_type == 'Liquid':
                worksheet.write('A38', 'BB', boldc)
                worksheet.write('A59', 'CC', boldc)
                worksheet.write('A60', '28', cell_format2)
                worksheet.write('A61', '29', cell_format2)
                worksheet.write('A62', '30', cell_format2)
                worksheet.write('A63', '31', cell_format2)
                worksheet.write('A64', '32', cell_format2)

                worksheet.merge_range("B38:M38", 'CV Plot', bold)
                worksheet.merge_range("B59:J59", 'Notes', bold)
                worksheet.merge_range("B60:J60", '', cell_format)
                worksheet.merge_range("B61:J61", '', cell_format)
                worksheet.merge_range("B62:J62", '', cell_format)
                worksheet.merge_range("B63:J63", '', cell_format)
                worksheet.merge_range("B64:J64", '', cell_format)


                worksheet.merge_range("K59:M59", 'Revision Control', bold)
                worksheet.write('K60', 'Rev.', cell_format)
                worksheet.write('L60', 'By.', cell_format)
                worksheet.write('M60', 'Date.', cell_format)

                worksheet.write('K61', '', cell_format2)
                worksheet.write('K62', '', cell_format2)
                worksheet.write('K63', '', cell_format2)
                worksheet.write('K64', '', cell_format2)

                worksheet.write('L61', '', cell_format2)
                worksheet.write('L62', '', cell_format2)
                worksheet.write('L63', '', cell_format2)
                worksheet.write('L64', '', cell_format2)

                worksheet.write('M61', '', cell_format2)
                worksheet.write('M62', '', cell_format2)
                worksheet.write('M63', '', cell_format2)
                worksheet.write('M64', '', cell_format2)

            



            
            
            count+=1
            f_cnt+=1
    workbook.close()
            

def createSpecSheet(case_data, units, other, act_):
    print(f'INSIDE FUNNCSSS {case_data}')
    workbook = xlsxwriter.Workbook('E:\Sizing_Reports\controlvalve_specsheet.xlsx')
    print(f'LENCASEDATA {len(case_data)}')
    for i in range(len(case_data)):
        print(f'INSIDE I {i}')
        current_datetime = datetime.today().date().timetuple()
        fluid_state = case_data[i][0][16]
        print(fluid_state)
        if fluid_state.lower() == 'liquid':
            data_valve = ['Specific Gravity', 'Kinematic Viscosity', 'Vapor Pressure', 'Liquid Pr. Recovery Factor']
        else:
            data_valve = ['Molecular Weight', 'Specific Heat Ratio', 'Inlet compressibility Factor',
                          'Pressure drop ratio factor']

        if other[i][21] == 'globe':
            data_trim = ['CV/Characteristic', 'Flow Direction/ Balancing', 'Stem Material', 'Plug Material', 'Seat Material', 'Cage or Clamp Material',"Percentage Opening", '%', 'm/s','Seat Leakage Class','Bellows Material']
            
            
        else:
            data_trim = ['CV/Characteristic', 'Flow Direction', 'Shaft Material', 'Disc Material', 'Seat Material', "Seat Leakage Class", "Degree of Opening", 'degree', 'mach','Bellows Material','']
            
        
        str_current_datetime = str(current_datetime)
        a__ = datetime.now()
        a_ = a__.strftime("%a, %d %b %Y %H-%M-%S")

        worksheet = workbook.add_worksheet(f"sheet-{int(i) + 1}")
        worksheet.set_paper(9)
        worksheet.fit_to_pages(2, 1)
        worksheet.print_area('A2:M72')
        worksheet.set_margins(0.36, 0.2, 0.4, 0.01)
        worksheet.set_header(margin=0.2)
        worksheet.set_footer(margin=0.28)

        bold = workbook.add_format(
            {'bold': True, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1, 'font': 'Arial', 'font_size': 9, 'border': 7, 'bg_color': '#EFF5F5'})
        bold1 = workbook.add_format(
            {'bold': True, 'bottom': 1, 'top': 1, 'left': 1, 'border': 7, 'bg_color': '#EFF5F5'})
        boldc = workbook.add_format(
            {'align': 'center', 'bold': True, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1,'font': 'Arial', 'font_size': 9, 'border': 7,
            'bg_color': '#EFF5F5'})
        br = workbook.add_format({'bottom': 1, 'top': 1, 'right': 1, 'border': 7})
        h1 = workbook.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'font': 'Arial', 'bottom': 1, 'top': 1, 'right': 1, 'left': 1})
        h2 = workbook.add_format({'bold': 0, 'font_size': 10, 'font': 'Arial'})
        h3 = workbook.add_format({'bold': 0, 'font_size': 9, 'font': 'Arial'})
        f1 = workbook.add_format({'bold': 0, 'font_size': 9, 'font': 'Arial', 'align': 'right'})

        cell_format1 = workbook.add_format({'bottom': 1, 'top': 1, 'font': 'Arial', 'font_size': 9, 'border': 7, 'right': 0, 'left': 0, 'align':'left'})
        cell_formatL = workbook.add_format({'bottom': 1, 'top': 1, 'font': 'Arial', 'font_size': 9,  'right': 1,'border': 7, 'left': 0 })
        cell_format = workbook.add_format(
            {'bold': 0, 'font_color': 'black', 'font': 'Arial', 'font_size': 9, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1,
            'border': 7,'align':'left'})

        cell_format2 = workbook.add_format(
            {'bold': 0, 'align': 'center', 'font_color': 'black', 'font': 'Arial', 'font_size': 9, 'bottom': 1, 'top': 1,
            'right': 1, 'left': 1, 'border': 7})

        cell_format3 = workbook.add_format(
        {'bold': True, 'align': 'center', 'font_color': 'black', 'font': 'Arial', 'font_size': 9, 'bottom': 1, 'top': 1,
            'right': 1, 'left': 1, 'border': 7})

        column_width = [{'name': 'A1:A73', 'size': 5}, {'name': 'B1:B73', 'size': 16.57}, {'name': 'C1:C73', 'size': 11},
                        {'name': 'D1:D73', 'size': 12},
                        {'name': 'E1:E73', 'size': 7.5}, {'name': 'F1:F73', 'size': 7}, {'name': 'G1:G73', 'size': 5},
                        {'name': 'H1:H73', 'size': 9.5},
                        {'name': 'I1:I73', 'size': 13}, {'name': 'J1:J73', 'size': 7.5}, {'name': 'K1:K73', 'size': 7},
                        {'name': 'L1:L73', 'size': 6.5}, {'name': 'M1:M73', 'size': 8}]

        for j in column_width:
            worksheet.set_column(j['name'], j['size'])

        # worksheet.set_column('A1:A73', 5)
        # worksheet.set_column('B1:B73', 24)
        # worksheet.set_column('C1:C73', 10)
        # worksheet.set_column('D1:D73', 12)
        # worksheet.set_column('E1:E73', 7)
        # worksheet.set_column('F1:F73', 5)
        # worksheet.set_column('G1:G73', 12)
        # worksheet.set_column('H1:H73', 12)
        # worksheet.set_column('I1:I73', 6)
        # worksheet.set_column('J1:J73', 6)
        # worksheet.set_column('K1:K73', 6)
        # worksheet.set_column('L1:L73', 6)

        # worksheet.insert_image('A3', 'https://fccommune.com/FCC_Sizing_Sheet_Logo.png', {'x_scale': 0.5, 'y_scale': 0.5})
        # worksheet.merge_range('A3:B7', '/newlogo.png',{'x_scale': 0.5, 'y_scale': 0.5})
        worksheet.merge_range('A3:B7', '', cell_format)  # Merge cells A3:B7

        # # Read the image file
        # with open('/static/images/newlogo.png', 'rb') as image_file:
        #     image_data = image_file.read()

        # Insert the image into cell A3
        worksheet.insert_image('A4', 'newlogo.png',{'x_scale': 0.16, 'y_scale': 0.16})

        worksheet.write('A1', '', )
        worksheet.merge_range('A2:M2', 'Control Valve Specification Sheet', cell_format3)
        worksheet.merge_range('A8:B8', "Application", cell_format)
        worksheet.write('A9:B9', 'Fluid State / Name', cell_format)
        worksheet.write('A10', 'AA', boldc)
        worksheet.write('A11', '1', cell_format2)
        worksheet.write('A12', '2', cell_format2)
        worksheet.write('A13', '3', cell_format2)
        worksheet.write('A14', '4', cell_format2)
        worksheet.write('A15', '5', cell_format2)
        worksheet.write('A16', '6', cell_format2)
        worksheet.write('A17', '7', cell_format2)
        worksheet.write('A18', '8', cell_format2)
        worksheet.write('A19', '9', cell_format2)
        worksheet.write('A20', '10', cell_format2)
        worksheet.write('A21', '11', cell_format2)
        worksheet.write('A22', '12', cell_format2)
        worksheet.write('A23', '13', cell_format2)
        worksheet.write('A24', '14', cell_format2)
        worksheet.write('A25', '15', cell_format2)
        # worksheet.write('A29', '')
        worksheet.write('A26', 'BB', boldc)
        worksheet.write('A27', '16', cell_format2)
        worksheet.write('A28', '17', cell_format2)
        worksheet.write('A29', '18', cell_format2)
        worksheet.write('A30', '19', cell_format2)
        worksheet.write('A31', '20', cell_format2)
        worksheet.write('A32', 'CC', boldc)
        worksheet.write('A33', '21', cell_format2)
        worksheet.write('A34', '22', cell_format2)
        worksheet.write('A35', '23', cell_format2)
        worksheet.write('A36', '24', cell_format2)
        worksheet.write('A37', '25', cell_format2)
        worksheet.write('A38', '26', cell_format2)
        worksheet.write('A39', '27', cell_format2)
        worksheet.write('A40', '28', cell_format2)
        worksheet.write('A41', '29', cell_format2)
        worksheet.write('A42', '30', cell_format2)
        worksheet.write('A43', '31', cell_format2)
        worksheet.write('A44', '32', cell_format2)
        worksheet.write('A45', '33', cell_format2)
        worksheet.write('A46', '34', cell_format2)
        worksheet.write('A47', 'DD', boldc)
        worksheet.write('A48', '35', cell_format2)
        worksheet.write('A49', '36', cell_format2)
        worksheet.write('A50', '37', cell_format2)
        worksheet.write('A51', '38', cell_format2)
        worksheet.write('A52', '39', cell_format2)
        worksheet.write('A53', '40', cell_format2)
        worksheet.write('A54', '41', cell_format2)
        worksheet.write('A55', '42', cell_format2)
        worksheet.write('A56', '43', cell_format2)
        worksheet.write('A57', 'HH', boldc)
        worksheet.write('A58', '71', cell_format2)
        worksheet.write('A59', '72', cell_format2)
        worksheet.write('A60', '73', cell_format2)
        worksheet.write('A61', '74', cell_format2)
        worksheet.write('A62', '75', cell_format2)
        worksheet.write('A63', '76', cell_format2)
        worksheet.write('A64', '77', cell_format2)
        worksheet.write('A65', '78', cell_format2)
        worksheet.write('A66', '79', cell_format2)
        worksheet.write('A67', '80', cell_format2)
        worksheet.write('A68', '81', cell_format2)
        worksheet.write('A69', '82', cell_format2)
        worksheet.write('A70', '83', cell_format2)
        worksheet.write('A71', '84', cell_format2)


        worksheet.write('B2', '')
        worksheet.write('B3', '')

        worksheet.merge_range("B10:C10", 'Valve Sizing Calculation', bold)
        worksheet.merge_range("B11:C11", 'Flow Rate', cell_format)
        worksheet.merge_range("B12:C12", 'Inlet Pressure', cell_format)
        worksheet.merge_range("B13:C13", 'Outlet Pressure', cell_format)
        worksheet.merge_range("B14:C14", 'Inlet Temperature', cell_format)
        worksheet.merge_range("B15:C15", data_valve[0], cell_format)
        worksheet.merge_range("B16:C16", data_valve[1], cell_format)
        worksheet.merge_range("B17:C17", data_valve[2], cell_format)
        worksheet.merge_range("B18:C18", data_valve[3], cell_format)
        worksheet.merge_range("B19:C19", 'Calculated Cv', cell_format)
        worksheet.merge_range("B20:C20", data_trim[6], cell_format)
        worksheet.merge_range("B21:C21", 'Noise level at 1m distance', cell_format)
        worksheet.merge_range("B22:C22", 'Inlet Pipe Velocity', cell_format)
        worksheet.merge_range("B23:C23", 'Outlet Valve Velocity', cell_format)
        # worksheet.merge_range("B24:C24", 'Trim Exit Velocity', cell_format)
        worksheet.merge_range("B24:C24", 'Warnings!!', cell_format)

        worksheet.merge_range("B26:F26", 'Pipe Line Data', bold)
        worksheet.merge_range("B27:C27", 'Size / Schedule Inlet', cell_format1)
        worksheet.merge_range("B28:C28", 'Size / Schedule Oulet', cell_format1)
        worksheet.merge_range("B29:C29", 'Insulation', cell_format1)
        worksheet.merge_range("B30:C30", 'Design Pressure', cell_format1)
        worksheet.merge_range("B31:C31", 'Design Temp. Min / Max', cell_format1)
        worksheet.merge_range("B32:F32", 'Valve Body & Bonnet', bold)
        worksheet.merge_range("B33:C33", 'Valve Mfg / Model No.', cell_format1)
        worksheet.merge_range("B34:C34", 'Size / Rating', cell_format1)
        worksheet.merge_range("B35:C35", 'Body Style', cell_format1)
        worksheet.merge_range("B36:C36", 'Body Material', cell_format1)
        worksheet.merge_range("B37:C37", 'End Connections', cell_format1)
        worksheet.merge_range("B38:C38", 'End Finish / Preparation', cell_format1)
        worksheet.merge_range("B39:C39", 'Body F/F Dim', cell_format1)
        worksheet.merge_range("B40:C40", 'Bonnet Type', cell_format1)
        worksheet.merge_range("B41:C41", 'Bonnet Material', cell_format1)
        worksheet.merge_range("B42:C42", 'Extension dimension', cell_format1)
        worksheet.merge_range("B43:C43", 'Body/Bonnet NDE1 / NDE2', cell_format1)
        worksheet.merge_range("B44:C44", 'Bonnet Bolting', cell_format1)
        worksheet.merge_range("B45:C45", 'Gasket Material', cell_format1)
        worksheet.merge_range("B46:C46", 'Gland Packing', cell_format1)

        worksheet.merge_range("B47:F47", 'Valve Trim', bold)

        worksheet.merge_range("B48:C48", 'Trim Size / Type', cell_format1)
        worksheet.merge_range("B49:C49", 'Rated CV / Characteristic', cell_format1)
        worksheet.merge_range("B50:C50", data_trim[1], cell_format1)
        worksheet.merge_range("B51:C51", data_trim[2], cell_format1)
        worksheet.merge_range("B52:C52", data_trim[3], cell_format1)
        worksheet.merge_range("B53:C53", data_trim[4], cell_format1)
        worksheet.merge_range("B54:C54", data_trim[5], cell_format1)
        worksheet.merge_range("B55:C55", data_trim[9], cell_format1)
        worksheet.merge_range("B56:C56", data_trim[10], cell_format1)
        # worksheet.merge_range("B56:J56", 'Bellows Material', cell_format1)


        worksheet.merge_range("B57:J57", 'Notes', bold)
        worksheet.merge_range("B59:J59", '', cell_format)
        worksheet.merge_range("B60:J60", '', cell_format)
        worksheet.merge_range("B61:J61", '', cell_format)
        worksheet.merge_range("B62:J62", '', cell_format)
        worksheet.merge_range("B63:J63", '', cell_format)
        worksheet.merge_range("B64:J64", '', cell_format)
        worksheet.merge_range("B65:J65", '', cell_format)
        worksheet.merge_range("B66:J66", '', cell_format)
        worksheet.merge_range("B67:J67", '', cell_format)
        worksheet.merge_range("B68:J68", '', cell_format)
        worksheet.merge_range("B69:J69", '', cell_format)
        worksheet.merge_range("B70:J70", '', cell_format)
        worksheet.merge_range("B71:J71", '', cell_format)
        worksheet.write('C1', '')
        # worksheet.write('C2', '')
        worksheet.write('C3', 'Customer', cell_format)
        worksheet.write('C4', 'Project', cell_format)
        worksheet.write('C5', 'End User', cell_format)
        worksheet.write('C6', 'RFQ No', cell_format)
        worksheet.write('C7', 'PO No', cell_format)
        worksheet.merge_range('C8:M8', '', cell_format)
        worksheet.merge_range('C9:F9', '', cell_format)

        # worksheet.write('D2', '')
        worksheet.merge_range('D3:H3', '', cell_format)
        worksheet.merge_range('D4:H4', '', cell_format)
        worksheet.merge_range('D5:H5', '', cell_format)
        worksheet.merge_range('D6:H6', '', cell_format)
        worksheet.merge_range('D7:H7', '', cell_format)

        worksheet.write('D10', 'Units', bold)
        worksheet.write('D11', units[i][0], cell_format)
        worksheet.write('D12', units[i][1], cell_format)
        worksheet.write('D13', units[i][2], cell_format)
        worksheet.write('D14', f'°{units[i][3]}', cell_format)
        worksheet.write('D15', '', cell_format)
        worksheet.write('D16', units[i][13], cell_format)
        worksheet.write('D17', units[i][12], cell_format)
        worksheet.write('D18', units[i][4], cell_format)
        worksheet.write('D19', units[i][5], cell_format)
        worksheet.write('D20', units[i][6], cell_format)
        worksheet.write('D21', units[i][7], cell_format)
        worksheet.write('D22', units[i][8], cell_format)
        worksheet.write('D23', units[i][9], cell_format)
        worksheet.write('D24', '', cell_format)
        worksheet.merge_range("D25:M25", '', cell_format)
        # worksheet.write('D26', '', cell_format)
        worksheet.write('D27', '', cell_format1)
        worksheet.write('D28', '', cell_format1)
        worksheet.write('D29', 'N/A', cell_format1)
        worksheet.write('D30', '', cell_format1)
        worksheet.write('D31', '', cell_format1)
        # worksheet.write('D32', '', cell_format1)
        worksheet.write('D33', 'FCC', cell_format1)
        worksheet.write('D34', '', cell_format1)
        worksheet.write('D35', '', cell_format1)
        worksheet.write('D36', '', cell_format1)
        worksheet.write('D37', '', cell_format1)
        worksheet.write('D38', '', cell_format1)
        worksheet.write('D39', 'N/A', cell_format1)
        worksheet.write('D40', '', cell_format1)
        worksheet.write('D41', '', cell_format1)
        worksheet.write('D42', '', cell_format1)
        worksheet.write('D43', '', cell_format1)
        worksheet.write('D44', '', cell_format1)
        worksheet.write('D45', '', cell_format1)
        worksheet.write('D46', '', cell_format1)
        # worksheet.write('D47', '', cell_format1)
        worksheet.write('D48', '', cell_format1)
        worksheet.write('D49', '', cell_format1)
        worksheet.write('D50', '', cell_format1)
        worksheet.write('D51', '', cell_format1)
        worksheet.write('D52', '', cell_format1)
        worksheet.write('D53', '', cell_format1)
        worksheet.write('D54', '', cell_format1)
        worksheet.write('D55', '', cell_format1)
        worksheet.write('D56', '', cell_format1)



        worksheet.write('E27', '', cell_format1)
        worksheet.write('E28', '', cell_format1)
        worksheet.write('E29', '', cell_format1)
        worksheet.write('E30', '', cell_format1)
        worksheet.write('E31', '', cell_format1)
        # worksheet.write('E32', '', cell_format1)
        worksheet.write('E33', '', cell_format1)
        worksheet.write('E34', '', cell_format1)
        worksheet.write('E35', '', cell_format1)
        worksheet.write('E36', '', cell_format1)
        worksheet.write('E37', '', cell_format1)
        worksheet.write('E38', '', cell_format1)
        worksheet.write('E39', '', cell_format1)
        worksheet.write('E40', '', cell_format1)
        worksheet.write('E41', '', cell_format1)
        worksheet.write('E42', '', cell_format1)
        worksheet.write('E43', '', cell_format1)
        worksheet.write('E44', '', cell_format1)
        worksheet.write('E45', '', cell_format1)
        worksheet.write('E46', '', cell_format1)
        # worksheet.write('E47', '', cell_format1)
        worksheet.write('E48', '', cell_format1)
        worksheet.write('E49', '', cell_format1)
        worksheet.write('E50', '', cell_format1)
        worksheet.write('E51', '', cell_format1)
        worksheet.write('E52', '', cell_format1)
        worksheet.write('E53', '', cell_format1)
        worksheet.write('E54', '', cell_format1)
        worksheet.write('E55', '', cell_format1)



        worksheet.merge_range("E10:F10", 'Case 1', bold)
        worksheet.merge_range("E11:F11", '', cell_format)
        worksheet.merge_range("E12:F12", '', cell_format)
        worksheet.merge_range("E13:F13", '', cell_format)
        worksheet.merge_range("E14:F14", '', cell_format)
        worksheet.merge_range("E15:F15", '', cell_format)
        worksheet.merge_range("E16:F16", '', cell_format)
        worksheet.merge_range("E17:F17", '', cell_format)
        worksheet.merge_range("E18:F18", '', cell_format)
        worksheet.merge_range("E19:F19", '', cell_format)
        worksheet.merge_range("E20:F20", '', cell_format)
        worksheet.merge_range("E21:F21", '', cell_format)
        worksheet.merge_range("E22:F22", '', cell_format)
        worksheet.merge_range("E23:F23", '', cell_format)
        worksheet.merge_range("E24:F24", '', cell_format)


        worksheet.write('F27', '', cell_format1)
        worksheet.write('F28', '', cell_format1)
        worksheet.write('F29', '', cell_format1)
        worksheet.write('F30', '', cell_format1)
        worksheet.write('F31', '', cell_format1)
        # worksheet.write('F32', '', cell_format1)
        worksheet.write('F33', '', cell_format1)
        worksheet.write('F34', '', cell_format1)
        worksheet.write('F35', '', cell_format1)
        worksheet.write('F36', '', cell_format1)
        worksheet.write('F37', '', cell_format1)
        worksheet.write('F38', '', cell_format1)
        worksheet.write('F39', '', cell_format1)
        worksheet.write('F40', '', cell_format1)
        worksheet.write('F41', '', cell_format1)
        worksheet.write('F42', '', cell_format1)
        worksheet.write('F43', '', cell_format1)
        worksheet.write('F44', '', cell_format1)
        worksheet.write('F45', '', cell_format1)
        worksheet.write('F46', '', cell_format1)
        # worksheet.write('F47', '', cell_format1)
        worksheet.write('F48', '', cell_format1)
        worksheet.write('F49', '', cell_format1)
        worksheet.write('F50', '', cell_format1)
        worksheet.write('F51', '', cell_format1)
        worksheet.write('F52', '', cell_format1)
        worksheet.write('F53', '', cell_format1)
        worksheet.write('F54', '', cell_format1)
        worksheet.write('F55', '', cell_format1)



        worksheet.write('G1', '', )
        worksheet.write('G2', '', )
        worksheet.write('G8', '', )
        worksheet.merge_range("G9:H9", 'Critical Pressure.', cell_format)
        worksheet.merge_range("G10:H10", 'Case 2', bold)
        worksheet.merge_range("G11:H11", '', cell_format)
        worksheet.merge_range("G12:H12", '', cell_format)
        worksheet.merge_range("G13:H13", '', cell_format)
        worksheet.merge_range("G14:H14", '', cell_format)
        worksheet.merge_range("G15:H15", '', cell_format)
        worksheet.merge_range("G16:H16", '', cell_format)
        worksheet.merge_range("G17:H17", '', cell_format)
        worksheet.merge_range("G18:H18", '', cell_format)
        worksheet.merge_range("G19:H19", '', cell_format)
        worksheet.merge_range("G20:H20", '', cell_format)
        worksheet.merge_range("G21:H21", '', cell_format)
        worksheet.merge_range("G22:H22", '', cell_format)
        worksheet.merge_range("G23:H23", '', cell_format)
        worksheet.merge_range("G24:H24", '', cell_format)
        # worksheet.write('G25', '',)
        worksheet.write('G26', 'EE', boldc)
        worksheet.write('G27', '44', cell_format2)
        worksheet.write('G28', '45', cell_format2)
        worksheet.write('G29', '46', cell_format2)
        worksheet.write('G30', '47', cell_format2)
        worksheet.write('G31', '48', cell_format2)
        worksheet.write('G32', '49', cell_format2)
        worksheet.write('G33', '50', cell_format2)
        worksheet.write('G34', '51', cell_format2)
        worksheet.write('G35', '52', cell_format2)
        worksheet.write('G36', '53', cell_format2)
        worksheet.write('G37', '54', cell_format2)
        worksheet.write('G38', 'FF', boldc)
        worksheet.write('G39', '55', cell_format2)
        worksheet.write('G40', '56', cell_format2)
        worksheet.write('G41', '57', cell_format2)
        worksheet.write('G42', '58', cell_format2)
        worksheet.write('G43', '59', cell_format2)
        worksheet.write('G44', '60', cell_format2)
        worksheet.write('G45', '61', cell_format2)
        worksheet.write('G46', '62', cell_format2)
        worksheet.write('G47', '63', cell_format2)
        worksheet.write('G48', '64', cell_format2)
        worksheet.write('G49', '65', cell_format2)
        worksheet.write('G50', '66', cell_format2)
        worksheet.write('G51', 'GG', boldc)
        worksheet.write('G52', '67', cell_format2)
        worksheet.write('G53', '68', cell_format2)
        worksheet.write('G54', '69', cell_format2)
        worksheet.write('G55', '70', cell_format2)
        worksheet.write('G56', '', cell_format2)


        worksheet.merge_range("H26:M26", 'Actuator Data', bold)
        worksheet.merge_range("H27:I27", "Actuator Type", cell_format1)
        worksheet.merge_range("H28:I28", "Actuator Model No. / Travel", cell_format1)
        worksheet.merge_range("H29:I29", "Actuator Size / Spring", cell_format1)
        worksheet.merge_range("H30:I30", "Air Fail Action", cell_format1)
        worksheet.merge_range("H31:I31", 'Air Supply / Set pressure', cell_format1)
        worksheet.merge_range("H32:I32", 'Actuator Orientation', cell_format1)
        worksheet.merge_range("H33:I33", 'Handwheel', cell_format1)
        worksheet.merge_range("H34:I34", 'Travel Stops', cell_format1)
        worksheet.merge_range("H35:I35", 'Tubing make / Size', cell_format1)
        worksheet.merge_range("H36:I36", 'Fittings', cell_format1)
        worksheet.merge_range("H37:I37", 'Open / Close / Trip Time', cell_format1)
        worksheet.merge_range("H38:M38", 'Accessories & Tubing', bold)
        worksheet.merge_range("H39:I39", 'Positioner Mfg / Model', cell_format1)
        worksheet.merge_range("H40:I40", 'Signal Action', cell_format1)
        worksheet.merge_range("H41:I41", 'AFR Mfg / Model', cell_format1)
        worksheet.merge_range("H42:I42", 'Position Transmitter', cell_format1)
        worksheet.merge_range("H43:I43", 'Limit switch / Proximity switch', cell_format1)
        worksheet.merge_range("H44:I44", 'Boosters', cell_format1)
        worksheet.merge_range("H45:I45", 'Pilot Valves', cell_format1)
        worksheet.merge_range("H46:I46", 'Air Lock Relay', cell_format1)
        worksheet.merge_range("H47:I47", 'I/P converter Mfg / Model', cell_format1)
        worksheet.merge_range("H48:I48", 'Solenoid Valve Mfg / Model', cell_format1)
        worksheet.merge_range("H49:I49", 'Solenoid Action', cell_format1)
        worksheet.merge_range("H50:I50", 'Volume Tank', cell_format1)
        worksheet.merge_range("H51:M51", 'Testing, Certification, Cleaning, Painting', bold)
        worksheet.merge_range("H52:I52", 'Testing', cell_format1)
        worksheet.merge_range("H53:I53", 'Assembly cleaning procedure', cell_format1)
        worksheet.merge_range("H54:I54", 'Painting Specification', cell_format1)
        worksheet.merge_range("H55:I55", 'Product Certification', cell_format1)

        worksheet.write('I1', '')
        worksheet.write('I2', '')
        worksheet.write('I3', 'Quote No', cell_format)
        worksheet.write('I4', 'W/O No', cell_format)
        worksheet.write('I5', 'Serial No', cell_format)
        worksheet.write('I6', 'Tag No', cell_format)
        worksheet.write('I7', 'Item No / Qty', cell_format)
        # worksheet.write('I8', '', cell_format)
        worksheet.write('I9', '', cell_format)
        worksheet.write('I10', 'Case 3', bold)
        worksheet.write('I11', '', cell_format)
        worksheet.write('I12', '', cell_format)
        worksheet.write('I13', '', cell_format)
        worksheet.write('I14', '', cell_format)
        worksheet.write('I15', '', cell_format)
        worksheet.write('I16', '', cell_format)
        worksheet.write('I17', '', cell_format)
        worksheet.write('I18', '', cell_format)
        worksheet.write('I19', '', cell_format)
        worksheet.write('I20', '', cell_format)
        worksheet.write('I21', '', cell_format)
        worksheet.write('I22', '', cell_format)
        worksheet.write('I23', '', cell_format)
        worksheet.write('I24', '', cell_format)
        # worksheet.write('I25', '', cell_format)

        worksheet.write('J27', '', cell_format1)
        # worksheet.write('J28', '', cell_format1)
        worksheet.write('J29', '', cell_format1)
        worksheet.write('J30', '', cell_format1)
        worksheet.write('J31', '', cell_format1)
        worksheet.write('J32', '', cell_format1)
        worksheet.write('J33', '', cell_format1)
        worksheet.write('J34', '', cell_format1)
        worksheet.write('J35', '', cell_format1)
        worksheet.write('J36', '', cell_format1)
        worksheet.write('J37', '', cell_format1)
        # worksheet.write('J38', '', cell_format1)
        worksheet.write('J39', '', cell_format1)
        worksheet.write('J40', '', cell_format1)
        worksheet.write('J41', '', cell_format1)
        worksheet.write('J42', '', cell_format1)
        worksheet.write('J43', '', cell_format1)
        worksheet.write('J44', '', cell_format1)
        worksheet.write('J45', '', cell_format1)
        worksheet.write('J46', '', cell_format1)
        worksheet.write('J47', '', cell_format1)
        worksheet.write('J48', '', cell_format1)
        worksheet.write('J49', '', cell_format1)
        worksheet.write('J50', '', cell_format1)
        # worksheet.write('J51', '', cell_format1)
        worksheet.write('J52', 'See Notes', cell_format1)
        worksheet.write('J53', act_[i]['cleaning'], cell_format1)
        worksheet.write('J54', act_[i]['paint_finish'], cell_format1)
        worksheet.write('J55', act_[i]['product_certs'], cell_format1)


        worksheet.write('K27', '', cell_format1)
        worksheet.write('K28', '', cell_format1)
        worksheet.write('K29', '', cell_format1)
        worksheet.write('K30', '', cell_format1)
        worksheet.write('K31', '', cell_format1)
        worksheet.write('K32', '', cell_format1)
        worksheet.write('K33', '', cell_format1)
        worksheet.write('K34', '', cell_format1)
        worksheet.write('K35', '', cell_format1)
        worksheet.write('K36', '', cell_format1)
        worksheet.write('K37', '', cell_format1)
        # worksheet.write('K38', '', cell_format1)
        worksheet.write('K39', '', cell_format1)
        worksheet.write('K40', '', cell_format1)
        worksheet.write('K41', '', cell_format1)
        worksheet.write('K42', '', cell_format1)
        worksheet.write('K43', '', cell_format1)
        worksheet.write('K44', '', cell_format1)
        worksheet.write('K45', '', cell_format1)
        worksheet.write('K46', '', cell_format1)
        worksheet.write('K47', '', cell_format1)
        worksheet.write('K48', '', cell_format1)
        worksheet.write('K49', '', cell_format1)
        worksheet.write('K50', '', cell_format1)
        # worksheet.write('K51', '', cell_format1)
        worksheet.write('K52', '', cell_format1)
        worksheet.write('K53', '', cell_format1)
        worksheet.write('K54', '', cell_format1)
        worksheet.write('K55', '', cell_format1)

        worksheet.merge_range("J3:M3", '', cell_format)
        worksheet.merge_range("J4:M4", '', cell_format)
        worksheet.merge_range("J5:M5", '', cell_format)
        worksheet.merge_range("J6:M6", '', cell_format)
        worksheet.merge_range("J7:M7", '', cell_format)
        worksheet.merge_range("J9:K9", 'Shutoff Pressure.', cell_format)
        worksheet.merge_range("J10:K10", 'Case 4', bold)
        worksheet.merge_range("J11:K11", '', cell_format)
        worksheet.merge_range("J12:K12", '', cell_format)
        worksheet.merge_range("J13:K13", '', cell_format)
        worksheet.merge_range("J14:K14", '', cell_format)
        worksheet.merge_range("J15:K15", '', cell_format)
        worksheet.merge_range("J16:K16", '', cell_format)
        worksheet.merge_range("J17:K17", '', cell_format)
        worksheet.merge_range("J18:K18", '', cell_format)
        worksheet.merge_range("J19:K19", '', cell_format)
        worksheet.merge_range("J20:K20", '', cell_format)
        worksheet.merge_range("J21:K21", '', cell_format)
        worksheet.merge_range("J22:K22", '', cell_format)
        worksheet.merge_range("J23:K23", '', cell_format)
        worksheet.merge_range("J24:K24", '', cell_format)
        worksheet.write('L27', '', cell_format1)
        worksheet.write('L28', '', cell_format1)
        worksheet.write('L29', '', cell_format1)
        worksheet.write('L30', '', cell_format1)
        worksheet.write('L31', '', cell_format1)
        worksheet.write('L32', '', cell_format1)
        worksheet.write('L33', '', cell_format1)
        worksheet.write('L34', '', cell_format1)
        worksheet.write('L35', '', cell_format1)
        worksheet.write('L36', '', cell_format1)
        worksheet.write('L37', '', cell_format1)
        # worksheet.write('L38', '', cell_format1)
        worksheet.write('L39', '', cell_format1)
        worksheet.write('L40', '', cell_format1)
        worksheet.write('L41', '', cell_format1)
        worksheet.write('L42', '', cell_format1)
        worksheet.write('L43', '', cell_format1)
        worksheet.write('L44', '', cell_format1)
        worksheet.write('L45', '', cell_format1)
        worksheet.write('L46', '', cell_format1)
        worksheet.write('L47', '', cell_format1)
        worksheet.write('L48', '', cell_format1)
        worksheet.write('L49', '', cell_format1)
        worksheet.write('L50', '', cell_format1)
        # worksheet.write('L51', '', cell_format1)
        worksheet.write('L52', '', cell_format1)
        worksheet.write('L53', '', cell_format1)
        worksheet.write('L54', '', cell_format1)
        worksheet.write('L55', '', cell_format1)

        worksheet.write('M27', '', cell_formatL)
        worksheet.write('M28', '', cell_formatL)
        worksheet.write('M29', '', cell_formatL)
        worksheet.write('M30', '', cell_formatL)
        worksheet.write('M31', '', cell_formatL)
        worksheet.write('M32', '', cell_formatL)
        worksheet.write('M33', '', cell_formatL)
        worksheet.write('M34', '', cell_formatL)
        worksheet.write('M35', '', cell_formatL)
        worksheet.write('M36', '', cell_formatL)
        worksheet.write('M37', '', cell_formatL)
        # worksheet.write('M38', '', cell_formatL)
        worksheet.write('M39', '', cell_formatL)
        worksheet.write('M40', '', cell_formatL)
        worksheet.write('M41', '', cell_formatL)
        worksheet.write('M42', '', cell_formatL)
        worksheet.write('M43', '', cell_formatL)
        worksheet.write('M44', '', cell_formatL)
        worksheet.write('M45', '', cell_formatL)
        worksheet.write('M46', '', cell_formatL)
        worksheet.write('M47', '', cell_formatL)
        worksheet.write('M48', '', cell_formatL)
        worksheet.write('M49', '', cell_formatL)
        worksheet.write('M50', '', cell_formatL)
        # worksheet.write('M51', '', cell_formatL)
        worksheet.write('M52', '', cell_formatL)
        worksheet.write('M53', '', cell_formatL)
        worksheet.write('M54', '', cell_formatL)
        worksheet.write('M55', '', cell_formatL)

        worksheet.merge_range("L9:M9", '', cell_format)
        worksheet.merge_range("L10:M10", 'Case 5', bold)
        worksheet.merge_range("L11:M11", '', cell_format)
        worksheet.merge_range("L12:M12", '', cell_format)
        worksheet.merge_range("L13:M13", '', cell_format)
        worksheet.merge_range("L14:M14", '', cell_format)
        worksheet.merge_range("L15:M15", '', cell_format)
        worksheet.merge_range("L16:M16", '', cell_format)
        worksheet.merge_range("L17:M17", '', cell_format)
        worksheet.merge_range("L18:M18", '', cell_format)
        worksheet.merge_range("L19:M19", '', cell_format)
        worksheet.merge_range("L20:M20", '', cell_format)
        worksheet.merge_range("L21:M21", '', cell_format)
        worksheet.merge_range("L22:M22", '', cell_format)
        worksheet.merge_range("L23:M23", '', cell_format)
        worksheet.merge_range("L24:M24", '', cell_format)
        

        worksheet.merge_range("K57:M57", 'Revision Control', bold)
        worksheet.write('J57', '', cell_format)
        worksheet.write('K58', 'Rev.', cell_format)

        # worksheet.write('J58', 'Rev.', cell_format)
        worksheet.write('J59', '', cell_format)
        worksheet.write('J60', '', cell_format)
        worksheet.write('J61', '', cell_format)
        worksheet.write('J62', '', cell_format)
        worksheet.write('J63', '', cell_format)
        worksheet.write('J64', '', cell_format)
        worksheet.write('J65', '', cell_format)
        worksheet.write('J66', '', cell_format)
        worksheet.write('J67', '', cell_format)
        worksheet.write('J68', '', cell_format)
        worksheet.write('J69', '', cell_format)
        worksheet.write('J70', '', cell_format)
        worksheet.write('J71', '', cell_format)

        # worksheet.write('K58', '', cell_format)
        worksheet.write('K59', '', cell_format)
        worksheet.write('K60', '', cell_format)
        worksheet.write('K61', '', cell_format)
        worksheet.write('K62', '', cell_format)
        worksheet.write('K63', '', cell_format)
        worksheet.write('K64', '', cell_format)
        worksheet.write('K65', '', cell_format)
        worksheet.write('K66', '', cell_format)
        worksheet.write('K67', '', cell_format)
        worksheet.write('K68', '', cell_format)
        worksheet.write('K69', '', cell_format)
        worksheet.write('K70', '', cell_format)
        worksheet.write('K71', '', cell_format)

        # worksheet.write('L57', '', cell_format)
        worksheet.write('L58', 'By', cell_format)
        worksheet.write('L59', '', cell_format)
        worksheet.write('L60', '', cell_format)
        worksheet.write('L61', '', cell_format)
        worksheet.write('L62', '', cell_format)
        worksheet.write('L63', '', cell_format)
        worksheet.write('L64', '', cell_format)
        worksheet.write('L65', '', cell_format)
        worksheet.write('L66', '', cell_format)
        worksheet.write('L67', '', cell_format)
        worksheet.write('L68', '', cell_format)
        worksheet.write('L69', '', cell_format)
        worksheet.write('L70', '', cell_format)
        worksheet.write('L71', '', cell_format)

        worksheet.write('M56', '', cell_formatL)
        worksheet.write('M7', '', cell_formatL)
        # worksheet.write('M57', '', cell_format)
        worksheet.write('M58', 'Date', cell_format)
        worksheet.write('M59', '', cell_format)
        worksheet.write('M60', '', cell_format)
        worksheet.write('M61', '', cell_format)
        worksheet.write('M62', '', cell_format)
        worksheet.write('M63', '', cell_format)
        worksheet.write('M64', '', cell_format)
        worksheet.write('M65', '', cell_format)
        worksheet.write('M66', '', cell_format)
        worksheet.write('M67', '', cell_format)
        worksheet.write('M68', '', cell_format)
        worksheet.write('M69', '', cell_format)
        worksheet.write('M70', '', cell_format)
        worksheet.write('M71', '', cell_format)

        worksheet.write('A72:B72', 'FLOW CONTROL COMMUNE', h3)
        worksheet.write('M72', 'FR/AE/004 Iss.01', f1)

        # add data
        
        worksheet.write('D3', str(case_data[i][0][24]), cell_format1)
        worksheet.write('D5', str(case_data[i][0][25]), cell_format1)
        worksheet.write('D6', str(case_data[i][0][26]), cell_format1)
        worksheet.write('D7', str(case_data[i][0][27]), cell_format1)
        worksheet.write('C9', f"{str(case_data[i][0][16])} / {other[i][30]}",cell_format1)
        worksheet.write('J3', str(case_data[i][0][22]), cell_format1)
        worksheet.write('J4', str(case_data[i][0][23]), cell_format1)
        worksheet.write('J6', str(case_data[i][0][14]), cell_format1)
        
        worksheet.write('D27', f"{str(case_data[i][0][18])} {units[i][10]}", cell_format1)
        worksheet.write('D28', f"{str(case_data[i][0][19])} {units[i][11]}", cell_format1)

        column_dict_tuple = [('E', 0), ('G', 1), ('I', 2), ('J', 3), ('L', 4)]
        len_cases = len(case_data[i])
        if len_cases > 5:
            len_cases = 5
        for k in column_dict_tuple[:len_cases]:
            for j in range(14):
                print(f'CASEDATAITEMSPEC {case_data[i][k[1]][j]}')
                worksheet.write(f'{k[0]}{j + 11}', case_data[i][k[1]][j], cell_format1)
        # for j in range(14):
        #     worksheet.write(f'D{j + 11}', units[i][j], cell_format1)
        worksheet.write(f'E27', case_data[i][0][28], cell_format1)
        worksheet.write(f'E28', case_data[i][0][29], cell_format1)

        # other[i] values input
        worksheet.write(f'J5', other[i][0], cell_format1)
        worksheet.write(f'H8', str(other[i][1]), cell_format1)
        worksheet.write(f'D4', str(other[i][2]), cell_format)
        worksheet.write(f'I9', f"{other[i][3]}",cell_format1)
        # worksheet.write(f'J11', other[i][4], f1)
        worksheet.write(f'L9', other[i][5],cell_format1)
        worksheet.write(f'D34', f"{other[i][6]} ",cell_format1)
        # worksheet.write(f'D38', other[i][7], cell_format1)
        worksheet.write(f'E34', f"/ {other[i][8]}",cell_format1)
        worksheet.write(f'D36', other[i][9], cell_format1)
        worksheet.write(f'D40', other[i][10], cell_format1)
        worksheet.write(f'D43', other[i][11], cell_format1)
        # worksheet.write(f'E43', other[i][12], cell_format1)
        worksheet.write(f'D45', other[i][13], cell_format1)
        worksheet.write(f'E48', f"/ {other[i][14]}", cell_format1)
        worksheet.write(f'D50', other[i][15], cell_format1)
        worksheet.write(f'D53', other[i][16], cell_format1)
        worksheet.write(f'D52', other[i][17], cell_format1)
        
        worksheet.write(f'D37', other[i][19], cell_format1)
        worksheet.write(f'D38', other[i][20], cell_format1)
        worksheet.write(f'D35', other[i][12], cell_format1)

        # other data
        worksheet.write(f'C34', case_data[i][0][1], cell_format1)
        worksheet.write(f'C35', case_data[i][0][3], cell_format1)
        worksheet.write(f'E33', f'/ {other[i][22]}', cell_format1)
        worksheet.write(f'D41', other[i][9], cell_format1)
        worksheet.write(f'D42', other[i][24], cell_format1)
        worksheet.write(f'D44', other[i][25], cell_format1)
        worksheet.write(f'D49', other[i][26], cell_format1)
        worksheet.write(f'E50', other[i][34], cell_format1)
        worksheet.write(f'C8', other[i][29], cell_format1)
        worksheet.write(f'D30', str(other[i][31]), cell_format1)
        worksheet.write(f'D31', str(other[i][33]), cell_format1)
        worksheet.write(f'E31', str(other[i][32]), cell_format1)
        worksheet.write(f'D46', str(other[i][36]), cell_format1)

        worksheet.write(f'D48', other[i][37], cell_format1)
       
        worksheet.write(f'D50', other[i][39], cell_format1)
        worksheet.write(f'E49', f"/ {other[i][40]}", cell_format1)
        worksheet.write(f'D51', other[i][41], cell_format1)
        if other[i][21] == 'globe':
            worksheet.write(f'D54', other[i][35], cell_format1) #cage
            worksheet.write(f'D55', other[i][18], cell_format1) #seatleakage 
            worksheet.write(f'D56', 'N/A', cell_format1) #bellows 
        else:
            worksheet.write(f'D54', other[i][18], cell_format1) #seatleakage
            worksheet.write(f'D55', 'N/A', cell_format1) #bellows
           
 
        worksheet.write(f'J7', other[i][43], cell_format1)
        

        # Adding Item Notes
        if len(other[i][42]) > 0:
            for note_index in range(len(other[i][42])):
                if other[i][42][note_index].notesNumber != "General":
                    worksheet.write(f'B{58 + note_index}', f"{other[i][42][note_index].notesNumber}-{other[i][42][note_index].content}", cell_format1) # item notes
                else:
                    worksheet.write(f'B{58 + note_index}', f"{other[i][42][note_index].content}", cell_format1) # item notes
                
        # worksheet.write(f'D48', other[i][37], cell_format1)

        # Actuator Data
        worksheet.write(f'J27', act_[i]['act_type'], cell_format1)
       
        worksheet.write(f'L28', f"/ {act_[i]['act_travel']}", cell_format1)
        worksheet.write(f'J29', act_[i]['act_size'], cell_format1)
        worksheet.write(f'J30', act_[i]["fail_action"], cell_format1)
        worksheet.write(f'J31', f'{act_[i]["air_supply_pressure"]}', cell_format1)
        worksheet.write(f'L31', f'/ {act_[i]["set_pressure"]}', cell_format1)
        worksheet.write(f'J32', act_[i]['orientation'], cell_format1)
        worksheet.write(f'J33', act_[i]['handwheel'], cell_format1)
        worksheet.write(f'J28',act_[i]['actmodelno'],cell_format1)
        
        worksheet.write(f'L29', f'/ {act_[i]['spring']}', cell_format1)
        worksheet.write(f'J34', act_[i]['travel_stops'], cell_format1)
        worksheet.write(f'J35', act_[i]['tubing'], cell_format1)
        worksheet.write(f'J36', act_[i]['fittings'], cell_format1)
        worksheet.write(f'J37', act_[i]['open'], cell_format1)
        worksheet.write(f'K37', act_[i]['close'], cell_format1)
        worksheet.write(f'L37', '/ N/A', cell_format1)
        worksheet.write(f'L35', '/ N/A', cell_format1)
        # worksheet.write(f'J39', act_[''], cell_format1)

        # Accessories data
        worksheet.write(f'J39', other[i][28][0], cell_format1)
        # worksheet.write(f'L39', other[i][28][1], cell_format1)
        worksheet.write(f'L39', "/ See N2", cell_format1)
        worksheet.write(f'L41', "/ See N3", cell_format1)
        worksheet.write(f'J40', other[i][28][2], cell_format1)
        worksheet.write(f'J41', other[i][28][3], cell_format1)
        # worksheet.write(f'L41', other[i][28][4], cell_format1)
        worksheet.write(f'J42', other[i][28][5], cell_format1)
        worksheet.write(f'J43', other[i][28][6], cell_format1)
        worksheet.write(f'L43', f"/ {other[i][28][7]}", cell_format1)
        worksheet.write(f'J44', other[i][28][8], cell_format1)
        worksheet.write(f'J45', other[i][28][9], cell_format1)
        worksheet.write(f'J46', other[i][28][10], cell_format1)
        worksheet.write(f'J47', other[i][28][11], cell_format1)
        worksheet.write(f'L47', f"/ {other[i][28][12]}", cell_format1)
        worksheet.write(f'J48', other[i][28][13], cell_format1)
        worksheet.write(f'L48', f"/ {other[i][28][14]}", cell_format1) 
        worksheet.write(f'J49', other[i][28][15], cell_format1)
        worksheet.write(f'J50', other[i][28][16], cell_format1)

        worksheet.set_row(24, options={'hidden': True})


    workbook.close()


act_dict = {'v_type': 0, 'trim_type': 0, 'Balancing': 0, 'fl_direction': 'over', 'v_size': 0, 'v_size_unit': 'mm',
            'Seat_Dia': 0,
            'seat_dia_unit': 'mm', 'unbalance_area': 0, 'unbalance_area_unit': 'mm',
            'Stem_size': 0, 'Stem_size_unit': 'mm', 'Travel': 0, 'travel_unit': 'mm', 'Packing_Friction': 0,
            'packing_friction_unit': 'mm', 'Seat_Load_Factor': 0, 'Additional_Factor': 0, 'P1': 0, 'p1_unit': 'mm',
            'P2': 0, 'p2_unit': 'mm',
            'delP_Shutoff': 0, 'delP_Shutoff_unit': 'mm', 'unbal_force': 0, 'Kn': 0, 'delP_flowing': 0, 'act_type': 0,
            'fail_action': 0, 'act_size': 0, 'act_size_unit': 'mm', 'act_travel': 0, 'act_travel_unit': 'mm',
            'eff_area': 0, 'eff_area_unit': 'mm',
            'sMin': 0, 'sMax': 0, 'spring_rate': 0, 'spring_windup': 0, 'max_spring_load': 0,
            'max_air_supply': 0,
            'set_pressure': 0, 'set_pressure_unit': 'mm', 'act_thrust_down': 0, 'act_thrust_up': 0, 'friction_band': 0,
            'req_handWheel_thrust': 0, 'max_thrust': 0,
            'v_thrust_close': 0, 'v_thrust_open': 0, 'seat_load': 0
            }

proj_details = {'customer': '', 'project': 1, 'quote': '', 'wo': '', 'sNo': '', 'tagNo': '', 'qty': 1, 'item_no': 1}


def createActSpecSheet(header_,valvedatas_,actdatas_,units_,accessories_,forces_):

    print(f"ACTUATOR___ SIZING {header_}")
    current_datetime = datetime.today().date().timetuple()
    str_current_datetime = str(current_datetime)
    a__ = datetime.now()
    a_ = a__.strftime("%a, %d %b %Y %H-%M-%S")

    workbook = xlsxwriter.Workbook(f'E:\Sizing_Reports\controlact_specsheet.xlsx')
    for i in range(len(header_)):
        worksheet = workbook.add_worksheet()



        worksheet.set_paper(9)
        worksheet.fit_to_pages(2, 1)
        worksheet.print_area('A2:L72')
        worksheet.set_margins(0.3, 0.25, 0.4, 0.01)
        # worksheet.setRowHeight(2, 50)
        # worksheet.set_row(9, 10)
        worksheet.set_header(margin=0.2)
        worksheet.set_footer(margin=0.28)


        bold = workbook.add_format(
            {'bold': True, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1, 'font': 'Arial', 'font_size': 9, 'border': 7, 'bg_color': '#EFF5F5'})
        bold1 = workbook.add_format(
            {'bold': True, 'bottom': 1, 'top': 1, 'left': 1, 'border': 7, 'bg_color': '#EFF5F5'})
        boldc = workbook.add_format(
            {'align': 'center', 'bold': True, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1,'font': 'Arial', 'font_size': 9, 'border': 7,
            'bg_color': '#EFF5F5'})
        br = workbook.add_format({'bottom': 1, 'top': 1, 'right': 1, 'border': 7})
        h1 = workbook.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'font': 'Arial', 'bottom': 1, 'top': 1, 'right': 1, 'left': 1})
        h2 = workbook.add_format({'bold': 0, 'font_size': 10, 'font': 'Arial'})
        h3 = workbook.add_format({'bold': 0, 'font_size': 9, 'font': 'Arial'})
        f1 = workbook.add_format({'bold': 0, 'font_size': 9, 'font': 'Arial', 'align': 'right'})

        cell_format1 = workbook.add_format({'bottom': 1, 'top': 1, 'font': 'Arial', 'font_size': 9, 'border': 7, 'right': 0, 'left': 0})
        cell_formatL = workbook.add_format({'bottom': 1, 'top': 1, 'font': 'Arial', 'font_size': 9,  'right': 1,'border': 7, 'left': 0 })
        cell_format = workbook.add_format(
            {'bold': 0, 'font_color': 'black', 'font': 'Arial', 'font_size': 9, 'bottom': 1, 'top': 1, 'right': 1, 'left': 1,
            'border': 7, 'align':'left'})

        cell_format2 = workbook.add_format( 
            {'bold': 0, 'align': 'center', 'font_color': 'black', 'font': 'Arial', 'font_size': 9, 'bottom': 1, 'top': 1,
            'right': 1, 'left': 1, 'border': 7})

        cell_format3 = workbook.add_format(
        {'bold': True, 'align': 'center', 'font_color': 'black', 'font': 'Arial', 'font_size': 12, 'bottom': 1, 'top': 1,
            'right': 1, 'left': 1, 'border': 7})
        
        print(f'TYPEACTSHEET {valvedatas_[i][0]}')
        if valvedatas_[i][0] in ['Globe Straight','Globe Angle']:
            valveHeadings = ['Valve Style','Trim Type','Rating','Balancing','Flow Direction','Valve Size','Seat Diameter','Unbalanced Area','Valve Stem Size','Valve Travel','Packing Friction','Seat Load Factor','Additional Friction','P1 Max','P2 Min','DelpShutoff','Unbalance Force','Fluid Negative Gradient','Differential Pressure','']
            valveFHeadings = ['Valve Thrust (close)','Valve Thrust (open)','Seat Load Force','','','']
            actHeadings = ['Actuator Type','Air Fail Action','Actuator Size','Actuator Travel','Diaphragm Effective Area','Lower Benchset','Upper Benchset','Spring Rate','Spring Windup','Max Spring Load','Max Air Supply','Set Pressure','Friction Band','Req.Handwheel Thrust','Max Top/Side Mount HW Thrust']
            actFHeadings = ['Actuator Thrust (close)','Actuator Thrust (open)','','','','']
        
        else:
            valveHeadings = ['Valve Style','Trim Type','Flow Direction','Valve Travel','Valve Size','Disc Diameter','Shaft Dia','DelpShutoff','C_usc','C_usv','P1 Max','P2 Min','Co-Eff of Bushing','Co-Eff Packing Friction','Delp Flowing','A','B','Unseating Torque','Packing Torque','Friction Torque']
            valveFHeadings = ['Break to Open','Run to Open','End to Open','Break to Close','Run to Close','End to Close']
            actHeadings = ['Actuator Type','Air Fail Action','Actuator Size','Actuator Travel','Diaphragm Effective Area','Spring Set','Max Air Supply','Set Pressure','Req. Handwheel Thrust','Max Top/Side Mount HW Torque','','','','','']
            actFHeadings = ['Spring Torque (Start)','Spring Torque (Mid)','Spring Torque (End)','Air Torque (Start)','Air Torque (Mid)','Air Torque (End)']
        

        # column_width = [{'name': 'A1:A73', 'size': 5}, {'name': 'B1:B73', 'size': 16.57}, {'name': 'C1:C73', 'size': 11},
        #                 {'name': 'D1:D73', 'size': 12},
        #                 {'name': 'E1:E73', 'size': 7.5}, {'name': 'F1:F73', 'size': 7}, {'name': 'G1:G73', 'size': 5},
        #                 {'name': 'H1:H73', 'size': 9.5},
        #                 {'name': 'I1:I73', 'size': 13}, {'name': 'J1:J73', 'size': 7.5}, {'name': 'K1:K73', 'size': 7},
        #                 {'name': 'L1:L73', 'size': 6.5}, {'name': 'M1:M73', 'size': 8}]

        # for i in column_width:
        #     worksheet.set_column(i['name'], i['size'])
    


        worksheet.set_column('A1:A73', 5)
        worksheet.set_column('B1:B73', 16.57)
        worksheet.set_column('C1:C73', 4.29)
        worksheet.set_column('D1:D73', 10.43)
        worksheet.set_column('E1:E73', 11)
        worksheet.set_column('F1:F73', 10.86)
        worksheet.set_column('G1:G73', 5)
        worksheet.set_column('H1:H73', 16.57)
        worksheet.set_column('I1:I73', 7.43)
        worksheet.set_column('J1:J73', 10.43)
        worksheet.set_column('K1:K73', 10)
        worksheet.set_column('L1:L73', 11)

        worksheet.insert_image('A4', 'newlogo.png', {'x_scale': 0.18, 'y_scale': 0.18})
        worksheet.merge_range('A3:B7', "", cell_format)

        worksheet.write('A1', '', )
        worksheet.merge_range('A2:L2', 'Actuator Sizing Calculation Sheet', cell_format3)

        worksheet.write('A10', 'AA', boldc)
        worksheet.write('A11', '1', cell_format2)
        worksheet.write('A12', '2', cell_format2)
        worksheet.write('A13', '3', cell_format2)   
        worksheet.write('A14', '4', cell_format2)
        worksheet.write('A15', '5', cell_format2)
        worksheet.write('A16', '6', cell_format2)
        worksheet.write('A17', '7', cell_format2)
        worksheet.write('A18', '8', cell_format2)
        worksheet.write('A19', '9', cell_format2)
        worksheet.write('A20', '10', cell_format2)
        worksheet.write('A21', '11', cell_format2)
        worksheet.write('A22', '12', cell_format2)
        worksheet.write('A23', '13', cell_format2)
        worksheet.write('A24', '14', cell_format2)
        worksheet.write('A25', '15', cell_format2)
        # worksheet.write('A29', '')
        worksheet.write('A26', '16', cell_format2)
        worksheet.write('A27', '17', cell_format2)
        worksheet.write('A28', '18', cell_format2)
        worksheet.write('A29', '19', cell_format2)
        worksheet.write('A30', '20', cell_format2)
        worksheet.write('A31', 'CC', boldc)
        worksheet.write('A32', '41', cell_format2)
        worksheet.write('A33', '42', cell_format2)
        worksheet.write('A34', '43', cell_format2)
        worksheet.write('A35', '44', cell_format2)
        worksheet.write('A36', '45', cell_format2)
        worksheet.write('A37', '46', cell_format2)
        worksheet.write('A38', 'EE', boldc)
        worksheet.write('A39', '53', cell_format2)
        worksheet.write('A40', '54', cell_format2)
        worksheet.write('A41', '55', cell_format2)
        worksheet.write('A42', '56', cell_format2)
        worksheet.write('A43', '57', cell_format2)
        worksheet.write('A44', '58', cell_format2)
        worksheet.write('A45', '59', cell_format2)
        worksheet.write('A46', '60', cell_format2)
        worksheet.write('A47', '61', cell_format2)
        worksheet.write('A48', '62', cell_format2)
        worksheet.write('A49', 'GG', boldc)
        worksheet.write('A50', '73', cell_format2)
        worksheet.write('A51', '74', cell_format2)
        worksheet.write('A52', '75', cell_format2)
        worksheet.write('A53', '76', cell_format2)
        worksheet.write('A54', '77', cell_format2)
        worksheet.write('A55', '78', cell_format2)
        worksheet.write('A56', '79', cell_format2)
        worksheet.write('A57', '80', cell_format2)
        worksheet.write('A58', '81', cell_format2)
        worksheet.write('A59', '82', cell_format2)
        worksheet.write('A60', '83', cell_format2)
        worksheet.write('A61', '84', cell_format2)
        worksheet.write('A62', '85', cell_format2)
        worksheet.write('A63', '86', cell_format2)
        worksheet.write('A64', '87', cell_format2)
        worksheet.write('A65', '88', cell_format2)
        worksheet.write('A66', '89', cell_format2)
        worksheet.write('A67', '90', cell_format2)
        worksheet.write('A68', '91', cell_format2)
        worksheet.write('A69', '92', cell_format2)
        worksheet.write('A70', '93', cell_format2)
        worksheet.write('A71', '94', cell_format2)


        worksheet.write('B2', '')
        worksheet.write('B3', '')

        worksheet.merge_range("B10:C10", 'Valve Data', bold)
        worksheet.merge_range("B11:C11", valveHeadings[0], cell_format)
        worksheet.merge_range("B12:C12", valveHeadings[1], cell_format)
        worksheet.merge_range("B13:C13", valveHeadings[2], cell_format)
        worksheet.merge_range("B14:C14", valveHeadings[3], cell_format)
        worksheet.merge_range("B15:C15", valveHeadings[4], cell_format)
        worksheet.merge_range("B16:C16", valveHeadings[5], cell_format)
        worksheet.merge_range("B17:C17", valveHeadings[6], cell_format)
        worksheet.merge_range("B18:C18", valveHeadings[7], cell_format)
        worksheet.merge_range("B19:C19", valveHeadings[8], cell_format)
        worksheet.merge_range("B20:C20", valveHeadings[9], cell_format)
        worksheet.merge_range("B21:C21", valveHeadings[10], cell_format)
        worksheet.merge_range("B22:C22", valveHeadings[11], cell_format)
        worksheet.merge_range("B23:C23", valveHeadings[12], cell_format)
        worksheet.merge_range("B24:C24", valveHeadings[13], cell_format)
        worksheet.merge_range("B25:C25", valveHeadings[14], cell_format)
        worksheet.merge_range("B26:C26", valveHeadings[15], cell_format)
        worksheet.merge_range("B27:C27", valveHeadings[16], cell_format1)
        worksheet.merge_range("B28:C28", valveHeadings[17], cell_format1)
        worksheet.merge_range("B29:C29", valveHeadings[18], cell_format1)
        worksheet.merge_range("B30:C30", valveHeadings[19], cell_format1)

        worksheet.merge_range("B31:C31", 'Valve Force', bold)
        worksheet.merge_range("B32:C32", valveFHeadings[0], cell_format1)
        worksheet.merge_range("B33:C33", valveFHeadings[1], cell_format1)
        worksheet.merge_range("B34:C34", valveFHeadings[2], cell_format1)
        worksheet.merge_range("B35:C35", valveFHeadings[3], cell_format1)
        worksheet.merge_range("B36:C36", valveFHeadings[4], cell_format1)
        worksheet.merge_range("B37:C37", valveFHeadings[5], cell_format1)

        worksheet.merge_range("B38:C38", 'Accessories', bold)
        worksheet.merge_range("B39:C39", 'Positioner', cell_format1)
        worksheet.merge_range("B40:C40", 'AFR', cell_format1)
        worksheet.merge_range("B41:C41", 'Position Transmitter', cell_format1)
        worksheet.merge_range("B42:C42", 'Limit / Proximity Switch', cell_format1)
        worksheet.merge_range("B43:C43", 'Booster', cell_format1)
        worksheet.merge_range("B44:C44", 'Pilot Valve', cell_format1)
        worksheet.merge_range("B45:C45", 'Airlock Relay', cell_format1)
        worksheet.merge_range("B46:C46", 'IP COnverter', cell_format1)
        worksheet.merge_range("B47:C47", 'Solenoid Valve', cell_format1)
        worksheet.merge_range("B48:C48", 'Volume Tank', cell_format1)


        worksheet.merge_range("B49:I49", 'Notes', bold)

        worksheet.merge_range("B50:I50", 'N1- MPI-Body/Bnt Castings/Forgings to ASME B16.34 at Foundry/Forge', cell_format)
        worksheet.merge_range("B51:I51", '', cell_format)
        worksheet.merge_range("B52:I52", '', cell_format)
        worksheet.merge_range("B53:I53", '', cell_format)
        worksheet.merge_range("B54:I54", '', cell_format)
        worksheet.merge_range("B55:I55", '', cell_format)
        worksheet.merge_range("B56:I56", '', cell_format)
        worksheet.merge_range("B57:I57", '', cell_format)
        worksheet.merge_range("B58:I58", '', cell_format)
        worksheet.merge_range("B59:I59", '', cell_format)
        worksheet.merge_range("B60:I60", '', cell_format)
        worksheet.merge_range("B61:I61", '', cell_format)
        worksheet.merge_range("B62:I62", '', cell_format)
        worksheet.merge_range("B63:I63", '', cell_format)
        worksheet.merge_range("B64:I64", '', cell_format)
        worksheet.merge_range("B65:I65", '', cell_format)
        worksheet.merge_range("B66:I66", '', cell_format)
        worksheet.merge_range("B67:I67", '', cell_format)
        worksheet.merge_range("B68:I68", '', cell_format)
        worksheet.merge_range("B69:I69", '', cell_format)
        worksheet.merge_range("B70:I70", '', cell_format)
        worksheet.merge_range("B71:I71", '', cell_format)

        # worksheet.write('D2', '')
        worksheet.merge_range('C3:D3', 'Customer', cell_format)
        worksheet.merge_range('C4:D4', 'Project', cell_format)
        worksheet.merge_range('C5:D5', 'End User', cell_format)
        worksheet.merge_range('C6:D6', 'RFQ No', cell_format)
        worksheet.merge_range('C7:D7', 'PO No', cell_format)

        worksheet.merge_range('A8:B8', 'Application', cell_format)
        worksheet.merge_range('A9:B9', 'Fluid State / Name', cell_format)

        worksheet.merge_range('C8:L8', header_[i][10], cell_format)
        worksheet.merge_range('C9:L9', header_[i][11], cell_format)

        worksheet.write('H3', 'Quote No', cell_format)
        worksheet.write('H4', 'W/O No', cell_format)
        worksheet.write('H5', 'Serial No', cell_format)
        worksheet.write('H6', 'Tag No', cell_format)
        worksheet.write('H7', 'Item No #/ Qty', cell_format)

        worksheet.merge_range('E3:G3', header_[i][0], cell_format)
        worksheet.merge_range('E4:G4', header_[i][1], cell_format)
        worksheet.merge_range('E5:G5', header_[i][2], cell_format)
        worksheet.merge_range('E6:G6', header_[i][3], cell_format)
        worksheet.merge_range('E7:G7', header_[i][4], cell_format)

        worksheet.merge_range('I3:L3', header_[i][5], cell_format)
        worksheet.merge_range('I4:L4', header_[i][6], cell_format)
        worksheet.merge_range('I5:L5', header_[i][7], cell_format)
        worksheet.merge_range('I6:L6', header_[i][8], cell_format)
        worksheet.merge_range('I7:L7', header_[i][9], cell_format)

        worksheet.write('D10', 'Units', bold)
        worksheet.write('D11', '', cell_format)
        worksheet.write('D12', '', cell_format)
        worksheet.write('D13', '', cell_format)
        worksheet.write('D14', f'°{units_[i][0]}', cell_format)
        worksheet.write('D15', units_[i][1], cell_format)
        worksheet.write('D16', units_[i][2], cell_format)
        worksheet.write('D17', units_[i][3], cell_format)
        worksheet.write('D18', units_[i][4], cell_format)
        worksheet.write('D19', units_[i][5], cell_format)
        worksheet.write('D20', units_[i][6], cell_format)
        worksheet.write('D21', units_[i][7], cell_format)
        worksheet.write('D22', units_[i][8], cell_format)
        worksheet.write('D23', units_[i][9], cell_format)
        worksheet.write('D24', units_[i][10], cell_format)
        worksheet.write('D25', units_[i][11], cell_format)
        worksheet.write('D26', units_[i][12], cell_format)
        worksheet.write('D27', units_[i][13], cell_format)
        worksheet.write('D28', units_[i][14], cell_format)
        worksheet.write('D29', units_[i][15], cell_format)
        worksheet.write('D30', units_[i][16], cell_format)
        worksheet.write('D31', 'Units', bold)
        worksheet.write('D32', units_[i][17], cell_format)
        worksheet.write('D33', units_[i][18], cell_format)
        worksheet.write('D34', units_[i][19], cell_format)
        worksheet.write('D35', units_[i][20], cell_format)
        worksheet.write('D36', units_[i][21], cell_format)
        worksheet.write('D37', units_[i][22], cell_format)
        worksheet.write('D38', 'Make', bold)
        worksheet.write('D39', accessories_[i][0], cell_format)
        worksheet.write('D40', accessories_[i][2], cell_format)
        worksheet.write('D41', accessories_[i][4], cell_format)
        worksheet.write('D42', '', cell_format)
        worksheet.write('D43', accessories_[i][6], cell_format)
        worksheet.write('D44', accessories_[i][7], cell_format)
        worksheet.write('D45', accessories_[i][8], cell_format)
        worksheet.write('D46', accessories_[i][9], cell_format)
        worksheet.write('D47', accessories_[i][10], cell_format)
        worksheet.write('D48', accessories_[i][11], cell_format)

        worksheet.merge_range("E10:F10", 'Values', bold)
        worksheet.merge_range("E11:F11", valvedatas_[i][0], cell_format)
        worksheet.merge_range("E12:F12", valvedatas_[i][1], cell_format)
        worksheet.merge_range("E13:F13", valvedatas_[i][2], cell_format)
        worksheet.merge_range("E14:F14", valvedatas_[i][3], cell_format)
        worksheet.merge_range("E15:F15", valvedatas_[i][4], cell_format)
        worksheet.merge_range("E16:F16", valvedatas_[i][5], cell_format)
        worksheet.merge_range("E17:F17", valvedatas_[i][6], cell_format)
        worksheet.merge_range("E18:F18", valvedatas_[i][7], cell_format)
        worksheet.merge_range("E19:F19", valvedatas_[i][8], cell_format)
       
        worksheet.merge_range("E20:F20", valvedatas_[i][9], cell_format)
        worksheet.merge_range("E21:F21", valvedatas_[i][10], cell_format)
        worksheet.merge_range("E22:F22", valvedatas_[i][11], cell_format)
        worksheet.merge_range("E23:F23", valvedatas_[i][12], cell_format)
        worksheet.merge_range("E24:F24", valvedatas_[i][13], cell_format)
        worksheet.merge_range("E25:F25", valvedatas_[i][14], cell_format)
        worksheet.merge_range("E26:F26", valvedatas_[i][15], cell_format)
        worksheet.merge_range("E27:F27", valvedatas_[i][16], cell_format)
        worksheet.merge_range("E28:F28", valvedatas_[i][17], cell_format)
        worksheet.merge_range("E29:F29", valvedatas_[i][18], cell_format)
        worksheet.merge_range("E30:F30", '', cell_format)
        worksheet.merge_range("E31:F31", 'Values', bold)
        worksheet.merge_range("E32:F32", forces_[i][0], cell_format)
        worksheet.merge_range("E33:F33", forces_[i][1], cell_format)
        worksheet.merge_range("E34:F34", forces_[i][2], cell_format)
        worksheet.merge_range("E35:F35", forces_[i][3], cell_format)
        worksheet.merge_range("E36:F36", forces_[i][4], cell_format)
        worksheet.merge_range("E37:F37", forces_[i][5], cell_format)
        worksheet.merge_range("E38:F38", 'Model No.', bold)
        worksheet.merge_range("E39:F39", accessories_[i][1], cell_format)
        worksheet.merge_range("E40:F40", accessories_[i][3], cell_format)
        worksheet.merge_range("E41:F41", '', cell_format)
        worksheet.merge_range("E42:F42", accessories_[i][5], cell_format)
        worksheet.merge_range("E43:F43", '', cell_format)
        worksheet.merge_range("E44:F44", '', cell_format)
        worksheet.merge_range("E45:F45", '', cell_format)
        worksheet.merge_range("E46:F46", '', cell_format)
        worksheet.merge_range("E47:F47", '', cell_format)
        worksheet.merge_range("E48:F48", '', cell_format)


        worksheet.write('G10', 'BB', boldc)
        worksheet.write('G11', '21', cell_format2)
        worksheet.write('G12', '22', cell_format2)
        worksheet.write('G13', '23', cell_format2)
        worksheet.write('G14', '24', cell_format2)
        worksheet.write('G15', '25', cell_format2)
        worksheet.write('G16', '26', cell_format2)
        worksheet.write('G17', '27', cell_format2)
        worksheet.write('G18', '28', cell_format2) 
        worksheet.write('G19', '29', cell_format2)
        worksheet.write('G20', '30', cell_format2)
        worksheet.write('G21', '31', cell_format2)
        worksheet.write('G22', '32', cell_format2)
        worksheet.write('G23', '33', cell_format2)
        worksheet.write('G24', '34', cell_format2)

        worksheet.write('G25', '35', cell_format2)
        worksheet.write('G26', '36', cell_format2)
        worksheet.write('G27', '37', cell_format2)
        worksheet.write('G28', '38', cell_format2)
        worksheet.write('G29', '39', cell_format2)
        worksheet.write('G30', '40', cell_format2)
        worksheet.write('G31', 'DD', boldc)
        worksheet.write('G32', '47', cell_format2)
        worksheet.write('G33', '48', cell_format2)
        worksheet.write('G34', '49', cell_format2)
        worksheet.write('G35', '50', cell_format2)
        worksheet.write('G36', '51', cell_format2)
        worksheet.write('G37', '52', cell_format2)
        worksheet.write('G38', 'FF', boldc)
        worksheet.write('G39', '63', cell_format2)
        worksheet.write('G40', '64', cell_format2)
        worksheet.write('G41', '65', cell_format2)
        worksheet.write('G42', '66', cell_format2)
        worksheet.write('G43', '67', cell_format2)
        worksheet.write('G44', '68', cell_format2)
        worksheet.write('G45', '69', cell_format2)
        worksheet.write('G46', '70', cell_format2)
        worksheet.write('G47', '71', cell_format2)
        worksheet.write('G48', '72', cell_format2)

        worksheet.merge_range("H10:I10", 'Actuator Data', bold)
        worksheet.merge_range("H11:I11", actHeadings[0], cell_format)
        worksheet.merge_range("H12:I12", actHeadings[1], cell_format)
        worksheet.merge_range("H13:I13", actHeadings[2], cell_format)
        worksheet.merge_range("H14:I14", actHeadings[3], cell_format)
        worksheet.merge_range("H15:I15", actHeadings[4], cell_format)
        worksheet.merge_range("H16:I16", actHeadings[5], cell_format)
        worksheet.merge_range("H17:I17", actHeadings[6], cell_format)
        worksheet.merge_range("H18:I18", actHeadings[7], cell_format)
        worksheet.merge_range("H19:I19", actHeadings[8], cell_format)
        worksheet.merge_range("H20:I20", actHeadings[9], cell_format)
        worksheet.merge_range("H21:I21", actHeadings[10], cell_format)
        worksheet.merge_range("H22:I22", actHeadings[11], cell_format)
        worksheet.merge_range("H23:I23", actHeadings[12], cell_format)
        worksheet.merge_range("H24:I24", actHeadings[13], cell_format)
        worksheet.merge_range("H25:I25", actHeadings[14], cell_format)
        worksheet.merge_range("H26:I26", '', cell_format)
        worksheet.merge_range("H27:I27", '', cell_format)
        worksheet.merge_range("H28:I28", '', cell_format)
        worksheet.merge_range("H29:I29", '', cell_format)
        worksheet.merge_range("H30:I30", '', cell_format)
        worksheet.merge_range("H31:I31", 'Actuator Force', bold)
        worksheet.merge_range("H32:I32", actFHeadings[0], cell_format)
        worksheet.merge_range("H33:I33", actFHeadings[1], cell_format)
        worksheet.merge_range("H34:I34", actFHeadings[2], cell_format)
        worksheet.merge_range("H35:I35", actFHeadings[3], cell_format)
        worksheet.merge_range("H36:I36", actFHeadings[4], cell_format)
        worksheet.merge_range("H37:I37", actFHeadings[5], cell_format)
        worksheet.merge_range("H38:I38", 'Description', bold)
        worksheet.merge_range("H39:I39", 'Valve Open Time', cell_format)
        worksheet.merge_range("H40:I40", 'Valve Close Time', cell_format)
        worksheet.merge_range("H41:I41", '', cell_format)
        worksheet.merge_range("H42:I42", '', cell_format)
        worksheet.merge_range("H43:I43", '', cell_format)
        worksheet.merge_range("H44:I44", '', cell_format)
        worksheet.merge_range("H45:I45", '', cell_format)
        worksheet.merge_range("H46:I46", '', cell_format)
        worksheet.merge_range("H47:I47", '', cell_format)
        worksheet.merge_range("H48:I48", '', cell_format)



        worksheet.write('J10', 'Units', bold)
        worksheet.write('J11', '', cell_format)
        worksheet.write('J12', '', cell_format)
        worksheet.write('J13', '', cell_format)
        worksheet.write('J14', units_[i][23], cell_format)
        worksheet.write('J15', units_[i][24], cell_format)
        worksheet.write('J16', units_[i][25], cell_format)
        worksheet.write('J17', units_[i][26], cell_format)
        worksheet.write('J18', units_[i][27], cell_format)
        worksheet.write('J19', units_[i][28], cell_format)
        worksheet.write('J20', units_[i][29], cell_format)
        worksheet.write('J21', units_[i][30], cell_format)
        worksheet.write('J22', units_[i][31], cell_format)
        worksheet.write('J23', units_[i][32], cell_format)
        worksheet.write('J24', units_[i][33], cell_format)
        worksheet.write('J25', units_[i][34], cell_format)
        worksheet.write('J26', '', cell_format)
        worksheet.write('J27', '', cell_format)
        worksheet.write('J28', '', cell_format)
        worksheet.write('J29', '', cell_format)
        worksheet.write('J30', '', cell_format)
        worksheet.write('J31', 'Units', bold)
        worksheet.write('J32', units_[i][35], cell_format)
        worksheet.write('J33', units_[i][36], cell_format)
        worksheet.write('J34', units_[i][37], cell_format)
        worksheet.write('J35', units_[i][38], cell_format)
        worksheet.write('J36', units_[i][39], cell_format)
        worksheet.write('J37', units_[i][40], cell_format)
        worksheet.write('J38', 'Units', bold)
        worksheet.write('J39', units_[i][41], cell_format)
        worksheet.write('J40', units_[i][42], cell_format)
        worksheet.write('J41', '', cell_format)
        worksheet.write('J42', '', cell_format)
        worksheet.write('J43', '', cell_format)
        worksheet.write('J44', '', cell_format)
        worksheet.write('J45', '', cell_format)
        worksheet.write('J46', '', cell_format)
        worksheet.write('J47', '', cell_format)
        worksheet.write('J48', '', cell_format)


        worksheet.write('J50', 'Rev', cell_format)
        worksheet.write('J51', '', cell_format)
        worksheet.write('J52', '', cell_format)
        worksheet.write('J53', '', cell_format)
        worksheet.write('J54', '', cell_format)
        worksheet.write('J55', '', cell_format)
        worksheet.write('J56', '', cell_format)
        worksheet.write('J57', '', cell_format)
        worksheet.write('J58', '', cell_format)
        worksheet.write('J59', '', cell_format)
        worksheet.write('J60', '', cell_format)
        worksheet.write('J61', '', cell_format)
        worksheet.write('J62', '', cell_format)
        worksheet.write('J63', '', cell_format)
        worksheet.write('J64', '', cell_format)
        worksheet.write('J65', '', cell_format)
        worksheet.write('J66', '', cell_format)
        worksheet.write('J67', '', cell_format)
        worksheet.write('J68', '', cell_format)
        worksheet.write('J69', '', cell_format)
        worksheet.write('J70', '', cell_format)
        worksheet.write('J71', '', cell_format)


        worksheet.merge_range("K10:L10", 'Values', bold)
        worksheet.merge_range("K11:L11", actdatas_[i][0], cell_format)
        worksheet.merge_range("K12:L12", actdatas_[i][1], cell_format)
        worksheet.merge_range("K13:L13", actdatas_[i][2], cell_format)
        worksheet.merge_range("K14:L14", actdatas_[i][3], cell_format)
        worksheet.merge_range("K15:L15", actdatas_[i][4], cell_format)
        worksheet.merge_range("K16:L16", actdatas_[i][5], cell_format)
        worksheet.merge_range("K17:L17", actdatas_[i][6], cell_format)
        worksheet.merge_range("K18:L18", actdatas_[i][7], cell_format)
        worksheet.merge_range("K19:L19", actdatas_[i][8], cell_format)
        worksheet.merge_range("K20:L20", actdatas_[i][9], cell_format)
        worksheet.merge_range("K21:L21", actdatas_[i][11], cell_format)
        worksheet.merge_range("K22:L22", actdatas_[i][10], cell_format)
        worksheet.merge_range("K23:L23", actdatas_[i][12], cell_format)
        worksheet.merge_range("K24:L24", actdatas_[i][13], cell_format)
        worksheet.merge_range("K25:L25", actdatas_[i][14], cell_format)
        worksheet.merge_range("K26:L26", '', cell_format)
        worksheet.merge_range("K27:L27", '', cell_format)
        worksheet.merge_range("K28:L28", '', cell_format)
        worksheet.merge_range("K29:L29", '', cell_format)
        worksheet.merge_range("K30:L30", '', cell_format)
        worksheet.merge_range("K31:L31", 'Values', bold)
        worksheet.merge_range("K32:L32", forces_[i][6], cell_format)
        worksheet.merge_range("K33:L33", forces_[i][7], cell_format)
        worksheet.merge_range("K34:L34", forces_[i][8], cell_format)
        worksheet.merge_range("K35:L35", forces_[i][9], cell_format)
        worksheet.merge_range("K36:L36", forces_[i][10], cell_format)
        worksheet.merge_range("K37:L37", forces_[i][11], cell_format)
        worksheet.merge_range("K38:L38", 'Stroke Time', bold)
        worksheet.merge_range("K39:L39", forces_[i][12], cell_format)
        worksheet.merge_range("K40:L40", forces_[i][13], cell_format)
        worksheet.merge_range("K41:L41", '', cell_format)
        worksheet.merge_range("K42:L42", '', cell_format)
        worksheet.merge_range("K43:L43", '', cell_format)
        worksheet.merge_range("K44:L44", '', cell_format)
        worksheet.merge_range("K45:L45", '', cell_format)
        worksheet.merge_range("K46:L46", '', cell_format)
        worksheet.merge_range("K47:L47", '', cell_format)
        worksheet.merge_range("K48:L48", '', cell_format)

        worksheet.merge_range("J49:L49", 'Revison Control', bold)

        worksheet.write('K50', 'By', cell_format)
        worksheet.write('K51', '', cell_format)
        worksheet.write('K52', '', cell_format)
        worksheet.write('K53', '', cell_format)
        worksheet.write('K54', '', cell_format)
        worksheet.write('K55', '', cell_format)
        worksheet.write('K56', '', cell_format)
        worksheet.write('K57', '', cell_format)
        worksheet.write('K58', '', cell_format)
        worksheet.write('K59', '', cell_format)
        worksheet.write('K60', '', cell_format)
        worksheet.write('K61', '', cell_format)
        worksheet.write('K62', '', cell_format)
        worksheet.write('K63', '', cell_format)
        worksheet.write('K64', '', cell_format)
        worksheet.write('K65', '', cell_format)
        worksheet.write('K66', '', cell_format)
        worksheet.write('K67', '', cell_format)
        worksheet.write('K68', '', cell_format)
        worksheet.write('K69', '', cell_format)
        worksheet.write('K70', '', cell_format)
        worksheet.write('K71', '', cell_format)


        worksheet.write('L50', 'Date', cell_format)
        worksheet.write('L51', '', cell_format)
        worksheet.write('L52', '', cell_format)
        worksheet.write('L53', '', cell_format)
        worksheet.write('L54', '', cell_format)
        worksheet.write('L55', '', cell_format)
        worksheet.write('L56', '', cell_format)
        worksheet.write('L57', '', cell_format)
        worksheet.write('L58', '', cell_format)
        worksheet.write('L59', '', cell_format)
        worksheet.write('L60', '', cell_format)
        worksheet.write('L61', '', cell_format)
        worksheet.write('L62', '', cell_format)
        worksheet.write('L63', '', cell_format)
        worksheet.write('L64', '', cell_format)
        worksheet.write('L65', '', cell_format)
        worksheet.write('L66', '', cell_format)
        worksheet.write('L67', '', cell_format)
        worksheet.write('L68', '', cell_format) 
        worksheet.write('L69', '', cell_format)
        worksheet.write('L70', '', cell_format)
        worksheet.write('L71', '', cell_format)

        # worksheet.merge_range("A72:I72", 'FLOW CONTROL COMMUNE', h3)
        worksheet.write('A72:B72', 'FLOW CONTROL COMMUNE', h3)
        worksheet.write('L72', 'FR/AE/004/B Iss.01', f1)

    workbook.close()
